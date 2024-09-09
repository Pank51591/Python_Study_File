import serial
import sys
import time
import serial.tools.list_ports
import _thread
import os
import tkinter
import openpyxl
import datetime
from tkinter import *
#import pandas
from tkinter import messagebox
from numbers import Number
from decimal import Decimal
import ctypes
from ctypes import c_int8
from ctypes import c_int16
from distutils.core import setup

CPBYTElength = (42)   #控制参数的字节数
CPLength = (36)       #控制参数的个数

SPBYTElength = (81)     #状态参数的字节数
SPLength = (69)

MAX_RECV_BYTE = (81 + 3 + 2)

#rxd
rxd_data = [0] * MAX_RECV_BYTE

#txd
txd_data = [0] * (CPBYTElength + 3 + 2)

entry = {}  
entryVar ={}
    
checkVar = {}
checkButton = {}

verif_ok = 0
keyNum = 0
keyDelay = 0
sendEnable = 0
commCnt = 0
lostCnt = 0
Reget = 0
rowIndex = 2
startTest = 0
PersistentParameterReset = 0
saveEnable = 0
saveCPEnable = 0
LoadCPEnable = 0
receivingnotok = 0
lost = 0
recoverCnt = 0
verificationEnable = 0
CRC = 0
CRC_rxd = 0
sendRequest = 0
editOnce = 0
sendRequestCnt = 0
offset = 0
startup = 0
cplength_offset = 0


#UI板发送的控制参数
TxdTabs = (
    'OvenPower1','TargetChamberTemper1','TargetProbeTemper1','CurrentProbeTemper1','TargetHumi1','HeatTubeMode1','BoilerPower1','MainHotFanPower1','EvaptrayPower1',\
    'MagnetronPower1','ChamberLight1','DoorLight1','Inletvalve1','DispalyFan1','LEDFan1','PowerFan1','CameraFan1','ExhaustFan1','CoolingValve1','WetBallValve1',\
    'Drain1','ResetAll1','CleanFunc1','CSICtrlMode1','PID_P1','PID_I1','PID_D1','PID_Ctrl1[0]','PID_Ctrl1[1]','PID_Ctrl1[2]','PID_Ctrl1[3]','PID_Ctrl1[4]','PID_Ctrl1[5]',\
    'PID_Ctrl1[6]','PID_Ctrl1[7]','PID_Ctrl1[8]'
)



RxdTabs = (
    'OvenPower','TargetChamberTemper','TargetProbeTemper','reserveCommand','TargetHumi','ChamberLigeht','DoorLight','UpperTubePower','RearTubePower','DownTubePower',\
    'BoilerPower','HeatFanRatedPower','EvapTrayPower','MagnetronPower','Inletvalve','DispalyFan','LEDFan','PowerFan','CameraFan','WetBall','TopNTCTemper','BottomNTCTemper','WetBallNTCTemper',\
    'BoilerNTCTemper','EvapTrayNTCTemper','Door','FreshTankState','DiscardTankState','FreshWaterLackState','DiscardWaterLackState','CleanFuncState','HWVersion','SWVersion',\
    'U_12V','InputFre','ModelSelect','UpperHeatTubeRatedPower','DownHeatTubeRatedPower','RearHeatTubeRatedPower','HotFanRatedPower','BoilerRatedPower','EvapTrayRatedPower',\
    'CSICtrlMode','PID_P','PID_I','PID_D','PID_Ctrl[0]','PID_Ctrl[1]','PID_Ctrl[2]','PID_Ctrl[3]','PID_Ctrl[4]','PID_Ctrl[5]','PID_Ctrl[6]','PID_Ctrl[7]','PID_Ctrl[8]',\
    'ErrState[0]','ErrState[1]','ErrState[2]','ErrState[3]','ErrState[4]','ErrState[5]','ErrState[6]','ErrState[7]','ErrState[8]','ErrState[9]','ErrState[10]','ErrState[11]',\
    'ErrState[12]','ErrState[13]'
)



def CRC_Calculate_rxd_data(bytes):
    global ss
    global CRC_rxd

    CRC = 0xFFFF
    #for i in range(0, MAX_RECV_BYTE - 2):
    for i in range(0, bytes - 2):
        CRC ^= (ss[i])
        for j in range(0, 8):
            if (CRC & 0x01):
                CRC= (CRC >> 1) ^ 0x8408
            else:
                CRC >>= 0x01
    CRC_rxd = (~CRC) & 0xffff
    print('CRC_rxd:',CRC_rxd)



def putButton(row_s, col_s, name, task):
    button ={}
    
    button[name] = tkinter.Button(top, text = name, command = task, width = 14) 
    button[name].grid(row = row_s, column = col_s)


def putLable(row_s, col_s, name):
    lable = {}
    lable[name] = Label(top, text = name, width = 14, wraplength = 0)
    lable[name].grid(row = row_s, column = col_s)


def putEntry(row_s, col_s, name, value, sta):
    global entryVar
    global entry

    entryVar[name] = StringVar()
    entry[name] = Entry(top, bd = 4, textvariable = entryVar[name], width = 14, state = sta)
    entry[name].grid(row = row_s, column = col_s)
    entryVar[name].set(value)
    entry[name].select_range(0, 256 )


def putCheckbutton(row_s, col_s, name):
    global checkButton
    global checkVar
    
    checkVar[name] = IntVar()
    checkButton[name] = Checkbutton(top, text = name, variable = checkVar[name], \
        onvalue=1, offvalue=0)
    checkButton[name].var = checkVar[name]
    checkButton[name].grid(row = row_s, column = col_s)

def SendTask():
    global sendEnable
    
    sendEnable = 1


def LoadCPTask():
    global verif_ok
    global LoadCPEnable
    global verificationEnable
    global sendEnable
    global loadModel
    
    #while True:
    if LoadCPEnable:   
        if LoadCPEnable == 3:
            loadModel = (int(entryVar['Software Model'].get()) & 0xff)

        '''  
        elif LoadCPEnable == 2:
            loadModelBuf = entryVar['Machine Model'].get()
            loadModel = 0
            tmp = 'Model2SoftModel.xlsx'   
            print(tmp)  
            try:
                wb_model = openpyxl.load_workbook(tmp)
                print(wb_model)
                sheet_model = wb_model.get_sheet_by_name('Models')
                print(sheet_model)
            except:
                messagebox.showinfo('Error',"No Model2SoftModel.xlsx!")
            find = 0
            for k in range(0, 15):
                 
                tmp2 = sheet_model.cell(row = 1, column = k + 2).value
                print(tmp2)
                print(len(tmp2))
               
                model = ''
                cnt = 0
                for j in range(0, len(tmp2)):
                    if tmp2[j] != '\n':
                        model += tmp2[j]
                        cnt += 1
                        if cnt == 8:
                            cnt = 0
                            if loadModelBuf == model:
                                print(loadModelBuf)
                                find = 1
                                loadModel = k + 1
                                break
                            model = ''
                if find:
                    break  
        '''                          
                
        if loadModel >= 1 and  loadModel <= 15:     #选择机型
            try:
                tmp = entryVar['CPFile'].get() + '.xlsx'   
                print(tmp)  
                wb_cp = openpyxl.load_workbook(tmp)
                print(wb_cp)
                sheet_cp = wb_cp.get_sheet_by_name('Control Flow Parameter')
                print(sheet_cp)
                for i in range(0, CPLength - cplength_offset -1): 
                    tmp = sheet_cp.cell(row = i + 2, column = 2 + loadModel - 1).value
                    #tmp1 = sheet_cp.cell(row = 77 + 2, column = 2 + loadModel - 1).value
                    #print(tmp1)
                    if tmp != None:
                        entryVar[TxdTabs[i]].set(tmp)
                    else:
                        messagebox.showinfo('Error',"Invalid Value in *.xlsx.")
                        LoadCPEnable = 0
                        #if sendEnable == 2:
                            #sendEnable = 0
                            #verificationEnable = 0
                        break
                if LoadCPEnable == 2 or LoadCPEnable == 3:
                    sendEnable = 2
                    verificationEnable = 20
                print("Upload Control Flow Parameter")
                if LoadCPEnable == 1:
                    messagebox.showinfo('Upload',"The Control Flow Parameter is upload.") 
                #time.sleep(5)
            except:
                print ("Error: can\'t find folder or file can't be uploaded")
                messagebox.showinfo('Error',"can\'t find folder or file can't be uploaded.")
                time.sleep(5)
        else:
            if LoadCPEnable == 2:
                messagebox.showinfo('Error',"No this Machine Model Numbers!")
            else :
                messagebox.showinfo('Error',"Software Model should be 1 ~ 15!")
        LoadCPEnable = 0
            #time.sleep(5)
        #verification()
        #if verif_ok == 0:
            #LoadCPEnable = 0
        #else:
            #verif_ok = 1 #test
        #time.sleep(0.1) 


def topTask():
    #putCheckbutton(0, 1, "Port Open")
    
    putEntry(0, 3, "SaveFile", "SaveData", 'normal')
    #putCheckbutton(0, 4, "Save")
    putButton(0, 4, "Save", SaveEnableTask)
    
    #putLable(0, 5, 'Send Cycle [s]')
    #putEntry(0, 6, "Send Cycle", 0.5, 'normal')
    #putCheckbutton(0, 7, "Auto Send")
    putButton(0, 5, "Send Once", SendTask)
    putCheckbutton(0, 6, "Edit")

    putButton(0, 7, "Reset Persistent", ResetPersistentParameterTask)
    #putButton(0, 9, "Start Test", StartTestTask)
    #putLable(0, 10, 'Test Interval[0.1s]')
       
    putEntry(0, 8, "CPFile", "Control Parameter", 'normal')
    putLable(0,9,'Send Speed:')
    putEntry(0, 10, "Send Speed", 1000, 'normal') 

    #putButton(0, 9, "Save CP", SaveCPEnableTask)  
    #putButton(0, 10, "Load CP", LoadCPEnableTask) 
    #putCheckbutton(0, 11, "Disable 15M")

    putCheckbutton(1, 0, "Receiving")
    #putButton(1, 1, "Auto Length", autoLengthTask)
    putButton(1, 5, "Default CP", defaultTask)
    putLable(1,6,'Model Number:')
    putEntry(1, 7, 'Software Model', 1, 'normal')   
    putButton(1, 8, "Programming_s", ProgrammingTask)


    # CP            
    i = 0
    j = 0
    num = 0
    putLable(9, 5, "******************")
    putLable(9, 7, "******************")
    putLable(9, 6,  "UI board data")                    
    for lTab in TxdTabs:
        putLable(10 + j, i * 2, lTab)
        putEntry(10 + j, i * 2 +  1, lTab, 0, 'norm')
        num = num + 1
        j = j + 1

        # change row at which column
        if j == 8:  #5行
            j =  0
            i = i + 1
        if num == CPLength:
            break

 
    # RUN Status    
    i = 0
    j = 0
    num = 0
    putLable(27, 5, "******************")
    putLable(27, 7, "******************")
    putLable(27, 6,  "Ctrl board data")                    
    for lTab in RxdTabs:
        putLable(28 + j, i * 2, lTab)
        putEntry(28 + j, i * 2 +  1, lTab, 0, 'readonly')
        num = num + 1
        j = j + 1

        if j == 12:
            j = 0
            i = i + 1
        if num == SPLength:
            break
    

#                  
def PortTask():
    global top
    global portOK
    global ser
    global Reget
    global comNow
    global chechstatus
    global lostCnt
    global lost
       
    
    putCheckbutton(0, 2, "Port Open")
    Reget = 0
    baudRate = '2400'
    
    while True: 
        if Reget <= 1:      
            try: 
                ports = list(serial.tools.list_ports.comports())    #调用 serial.tools.list_ports.comports() 函数来列出所有可用的串行端口。
                    
            except:
                print ("Port dosen't exist")
                messagebox.showinfo( "!", "Port dosen't exist")
                #sys.exit() 
            portOK = 0
            mylist = [0,0,0,0]  
            j = 0 
            ok = 0
            for port, b, c, in ports:
                ok = 1
                print(ok)
                if port:
                    if Reget == 0:
                        varBaud = StringVar(top)
                        baudMenu = OptionMenu(top, varBaud, '300', '600', '1200', '2400', '4800', '9600', '14400','115200')
                        baudMenu.grid(row = 0, column = 1)                               
                        baudRate = '115200'
                    
                    varBaud.set(baudRate) 
                    print(baudRate)                        
                        
                    mylist[j] = port
                    var = StringVar(top)
                    if j == 0:
                        mymenu = OptionMenu(top, var, mylist[0])
                    elif j == 1:
                        mymenu = OptionMenu(top, var, mylist[0], mylist[1]) 
                    elif j == 2:
                        mymenu = OptionMenu(top, var, mylist[0], mylist[1], mylist[2])  
                    else:
                        mymenu = OptionMenu(top, var, mylist[0], mylist[1], mylist[2], mylist[3])             
                    j = j + 1    
                    mymenu.grid(row = 0, column = 0)
                    if Reget == 0:
                        comNow = mylist[0]
                    Reget = 2
                    var.set(comNow)
                    print(mylist)                 
         
        
        chechstatus = checkButton["Port Open"].var.get()
  
        if chechstatus: 
            if portOK == 0 and ok:   
                comNow = var.get()
                baudRate = varBaud.get()               
                try:
                    ser = serial.Serial(var.get(), baudRate, timeout = 2)  #0.2
                    print(ser)
                    portOK = 1
                    rxdtd = _thread.start_new_thread(rxdTask, ())
                    txdth = _thread.start_new_thread(txdTask, ())
                except:
                    print ("Port dosen't exist or be used by other application2!")
                    #if lostCnt < 6:
                    if lost == 0:
                        messagebox.showinfo( "!", "Port dosen't exist or be used by other application2") 
                        time.sleep(5)
                lost = 0
        else:
            if portOK:
                #rxdtd.join()
                #txdth.join()                
                ser.close()
                portOK = 0
                Reget = 1

    UIsend_Speed = 0
    UIsend_Speed = int(entryVar['Send Speed'].get())             
    time.sleep(UIsend_Speed * 0.001)    #控制发送的速度

def ProgrammingTask():
    global sendEnable
    global disable15MSend
    global LoadCPEnable
    global verificationEnable
    global editOnce
    
    #checkButton["Disable 15M"].deselect()
    checkButton["Edit"].select()
    while checkButton["Edit"].var.get() == 0:   # or checkButton["Disable 15M"].var.get():
        time.sleep(1)
    LoadCPEnable = 3
    LoadCPTask()
    #sendEnable = 0
    #time.sleep(1)
    #ResetPersistentParameterTask()
    #disable15MSend = 1 

    #sendEnable = 2
    #time.sleep(1)
    #checkButton["Edit"].deselect()
    #verificationEnable = 1
    print("verificationEnable:", verificationEnable)


def ResetPersistentParameterTask():
    global PersistentParameterReset
    
    PersistentParameterReset = 1

def SaveEnableTask():
    global saveEnable
    saveEnable = 1

def SaveCPEnableTask():
    global saveCPEnable
    saveCPEnable = 1

def LoadCPEnableTask(): 
    global LoadCPEnable
    
    checkButton["Edit"].select()
    LoadCPEnable = 1
    LoadCPTask()

#
def defaultTask():
    global sendEnable
    global disable15MSend
    global LoadCPEnable
    global verificationEnable
    global sendRequest
    global sendRequestCnt 
    
    #checkButton["Disable 15M"].deselect()
    #checkButton["Edit"].select()
    #while checkButton["Edit"].var.get() == 0:# or checkButton["Disable 15M"].var.get():
    #time.sleep(1)
    
    #time.sleep(1)
    #ResetPersistentParameterTask()
    #disable15MSend = 1   
    #time.sleep(1)
    #sendRequest = 0
    #sendRequestCnt = 0    
    sendEnable = 3


def rxdTask():
    global ss
    global ser
    global commCnt
    global rowIndex
    global lostCnt
    global receivingnotok
    global verificationEnable
    global CRC_rxd
    global offset
    global startup
    global cplength_offset
	
    time.sleep(0.3)

    while portOK:       
        try: 
            ss = ser.read(MAX_RECV_BYTE - offset)
        except:
            ss = b''    #空字符串
            #if lostCnt < 6:
                #ser.close()
                #lostCnt = 6
            pass

        #time.sleep(1)
        #print('quit rxdTask1 \r\n')
        #c = checkButton["edit"].var.get()
        c = checkButton["Edit"].var.get()
        ser.reset_input_buffer()  #清空缓冲区
        if  c == 0 and ss != b'':
            ss2 = ss[3:(len(ss) - 1)]
        
        #checksum = 0 
        #for i in range(len(ss) - 1):
            #checksum = checksum + ss[i]
        #checksum = checksum & 0xff
        
            CRC_Calculate_rxd_data(len(ss))
            #print(len(ss))

            checksum_read = int(ss[(len(ss) - 2)] << 8 | ss[(len(ss) - 1)])
        
            #print(CRC_rxd)
            #print(checksum_read)
            print(ss)
            #print(len(ss))
            #print(len(ss2))
            #print(ss[0])
            #print(ss[(len(ss) - 1)])
            #print(checksum)

            receivingnotok = 1
            if CRC_rxd == checksum_read:# and len(ss) == MAX_RECV_BYTE: 
                #print("App rxd:")
                startup = 1
                if offset == 0:
                    cplength_offset = 0
                elif offset == 3:
                    cplength_offset = 2

                #command = ss[1]
                #print(command)  
                checkButton["Receiving"].toggle()
                #print("checksum pass")
                #print(ss[67])
                #print(ss[68])
                lostCnt = 0
                receivingnotok = 0
                for columnIndex in range(0, len(ss2)):
                    rxd_data[columnIndex] = ss2[columnIndex]                  
                
                    
                ''' 
                # row = 1, item
                columnIndex = 1
                for item in RxdTabs:
                    sheet.cell(row = 1, column = columnIndex).value = item
                    columnIndex = columnIndex + 1
                
                # column = 1, time    
                sheet.cell(row = rowIndex, column = 1).value = datetime.datetime.now()    
                '''

                entryVar['OvenPower'].set((ss2[0]))
                entryVar['TargetChamberTemper'].set((ss2[2]<<8 | ss2[1]))
                entryVar['TargetProbeTemper'].set((ss2[4]<<8 | ss2[3]))
                entryVar['reserveCommand'].set((ss2[5]))
                entryVar['TargetHumi'].set((ss2[6]))
                entryVar['ChamberLigeht'].set((ss2[7]))
                entryVar['DoorLight'].set((ss2[8]))
                entryVar['UpperTubePower'].set((ss2[9]))
                entryVar['RearTubePower'].set((ss2[10]))
                entryVar['DownTubePower'].set((ss2[11]))
                entryVar['BoilerPower'].set((ss2[12]))
                entryVar['HeatFanRatedPower'].set((ss2[13]))
                entryVar['EvapTrayPower'].set((ss2[14]))
                entryVar['MagnetronPower'].set((ss2[15]))
                entryVar['Inletvalve'].set((ss2[16]))
                entryVar['DispalyFan'].set((ss2[17]))
                entryVar['LEDFan'].set((ss2[18]))
                entryVar['PowerFan'].set((ss2[19]))
                entryVar['CameraFan'].set((ss2[20]))
                entryVar['WetBall'].set((ss2[21]))
                entryVar['TopNTCTemper'].set((ss2[23]<<8 | ss2[22]))
                entryVar['BottomNTCTemper'].set((ss2[25]<<8 | ss2[24]))
                entryVar['WetBallNTCTemper'].set((ss2[27]<<8 | ss2[26]))
                entryVar['BoilerNTCTemper'].set((ss2[29]<<8 | ss2[28]))
                entryVar['EvapTrayNTCTemper'].set((ss2[31]<<8 | ss2[30]))
                entryVar['Door'].set((ss2[32]))
                entryVar['FreshTankState'].set((ss2[33]))
                entryVar['DiscardTankState'].set((ss2[34]))
                entryVar['FreshWaterLackState'].set((ss2[35]))
                entryVar['DiscardWaterLackState'].set((ss2[36]))
                entryVar['CleanFuncState'].set((ss2[37]))
                entryVar['HWVersion'].set((ss2[38]))
                entryVar['SWVersion'].set((ss2[40]<<8 | ss2[39]))
                entryVar['U_12V'].set((ss2[42]<<8 | ss2[41]))
                entryVar['InputFre'].set((ss2[43]))
                entryVar['ModelSelect'].set((ss2[44]))
                entryVar['UpperHeatTubeRatedPower'].set((ss2[45]))
                entryVar['DownHeatTubeRatedPower'].set((ss2[46]))
                entryVar['RearHeatTubeRatedPower'].set((ss2[47]))
                entryVar['HotFanRatedPower'].set((ss2[48]))
                entryVar['BoilerRatedPower'].set((ss2[49]))
                entryVar['EvapTrayRatedPower'].set((ss2[50]))
                entryVar['CSICtrlMode'].set((ss2[51]))
                entryVar['PID_P'].set((ss2[53]<<8 | ss2[52]))
                entryVar['PID_I'].set((ss2[55]<<8 | ss2[54]))
                entryVar['PID_D'].set((ss2[57]<<8 | ss2[56]))
                entryVar['PID_Ctrl[0]'].set((ss2[58]))
                entryVar['PID_Ctrl[1]'].set((ss2[59]))
                entryVar['PID_Ctrl[2]'].set((ss2[60]))
                entryVar['PID_Ctrl[3]'].set((ss2[61]))
                entryVar['PID_Ctrl[4]'].set((ss2[62]))
                entryVar['PID_Ctrl[5]'].set((ss2[63]))
                entryVar['PID_Ctrl[6]'].set((ss2[64]))
                entryVar['PID_Ctrl[7]'].set((ss2[65]))
                entryVar['PID_Ctrl[8]'].set((ss2[66]))
                entryVar['ErrState[0]'].set((ss2[67]))
                entryVar['ErrState[1]'].set((ss2[68]))
                entryVar['ErrState[2]'].set((ss2[69]))
                entryVar['ErrState[3]'].set((ss2[70]))
                entryVar['ErrState[4]'].set((ss2[71]))
                entryVar['ErrState[5]'].set((ss2[72]))
                entryVar['ErrState[6]'].set((ss2[73]))
                entryVar['ErrState[7]'].set((ss2[74]))
                entryVar['ErrState[8]'].set((ss2[75]))
                entryVar['ErrState[9]'].set((ss2[76]))
                entryVar['ErrState[10]'].set((ss2[77]))
                entryVar['ErrState[12]'].set((ss2[78]))
                entryVar['ErrState[13]'].set((ss2[79]))

                if verificationEnable == 1:
                    verificationEnable = 0
                    #verificationTask()    #对接收到的数据进行验证
        time.sleep(0.001) 


def txdTask():
    global sendEnable
    global disable15MSend
    global PersistentParameterReset
    global keyNum
    global keyDelay
    global ser
    global portOK
    global rowIndex
    global sheet_cp    
    global startTest
    global recoverCnt
    global CRC_txd
    global sendRequest
    global sendRequestCnt
    global offset
    global startup
    
    startup = 0
    cnt = 0
    time.sleep(0.5)
    while portOK:
        delay = 0
        first = 0
        
        '''
        c = startTest
        try:
            #if c:
            maxByte = CPBYTElength + 4 # no use
            #else:
            #     maxByte = 4
        except:
            pass
            messagebox.showinfo('error', 'not a int number!')
            time.sleep(5)
        
        if c:
            try:
                d = int(entryVar['Send Speed'].get())
            except:
                pass
                messagebox.showinfo('error', 'not a number!')
                time.sleep(5)  
        '''

        offset = 0
        txd_data[0] = 0xe3
        txd_data[1] = 1+2+CPBYTElength   #帧长度
        txd_data[2] = 0x00  #方向
        txd_data[3] = (int(entryVar['OvenPower1'].get()) & 0xff)
        txd_data[4] = (int(entryVar['TargetChamberTemper1'].get()) & 0xff)
        txd_data[5] = (int(entryVar['TargetChamberTemper1'].get()) >> 8) & 0xff
        txd_data[6] = (int(entryVar['TargetProbeTemper1'].get()) & 0xff)
        txd_data[7] = (int(entryVar['TargetProbeTemper1'].get()) >> 8 & 0xff)
        txd_data[8] = (int(entryVar['CurrentProbeTemper1'].get()) & 0xff)
        txd_data[9] = (int(entryVar['CurrentProbeTemper1'].get()) >> 8 & 0xff)
        txd_data[10] = (int(entryVar['TargetHumi1'].get()) & 0xff)
        txd_data[11] = (int(entryVar['TargetHumi1'].get()) >> 8 & 0xff)
        txd_data[12] = (int(entryVar['HeatTubeMode1'].get()) & 0xff)
        txd_data[13] = (int(entryVar['BoilerPower1'].get()) & 0xff)
        txd_data[14] = (int(entryVar['MainHotFanPower1'].get()) & 0xff)
        txd_data[15] = (int(entryVar['EvaptrayPower1'].get()) & 0xff)
        txd_data[16] = (int(entryVar['MagnetronPower1'].get()) & 0xff)
        txd_data[17] = (int(entryVar['ChamberLight1'].get()) & 0xff)
        txd_data[18] = (int(entryVar['DoorLight1'].get()) & 0xff)
        txd_data[19] = (int(entryVar['Inletvalve1'].get()) & 0xff)
        txd_data[20] = (int(entryVar['DispalyFan1'].get()) & 0xff)
        txd_data[21] = (int(entryVar['LEDFan1'].get()) & 0xff)
        txd_data[22] = (int(entryVar['PowerFan1'].get()) & 0xff)
        txd_data[23] = (int(entryVar['CameraFan1'].get()) & 0xff)
        txd_data[24] = (int(entryVar['ExhaustFan1'].get()) & 0xff)
        txd_data[25] = (int(entryVar['CoolingValve1'].get()) & 0xff)
        txd_data[26] = (int(entryVar['WetBallValve1'].get()) & 0xff)
        txd_data[27] = (int(entryVar['Drain1'].get()) & 0xff)
        txd_data[28] = (int(entryVar['ResetAll1'].get()) & 0xff)
        txd_data[29] = (int(entryVar['CleanFunc1'].get()) & 0xff)
        txd_data[30] = (int(entryVar['CSICtrlMode1'].get()) & 0xff)
        txd_data[31] = (int(entryVar['PID_P1'].get()) & 0xff)
        txd_data[32] = (int(entryVar['PID_P1'].get()) >> 8 & 0xff)
        txd_data[33] = (int(entryVar['PID_I1'].get()) & 0xff)
        txd_data[34] = (int(entryVar['PID_I1'].get()) >> 8 & 0xff)
        txd_data[35] = (int(entryVar['PID_D1'].get()) & 0xff)
        txd_data[36] = (int(entryVar['PID_D1'].get()) >> 8 & 0xff)
        txd_data[37] = (int(entryVar['PID_Ctrl1[0]'].get()) & 0xff)
        txd_data[38] = (int(entryVar['PID_Ctrl1[1]'].get()) & 0xff)
        txd_data[39] = (int(entryVar['PID_Ctrl1[2]'].get()) & 0xff)
        txd_data[40] = (int(entryVar['PID_Ctrl1[3]'].get()) & 0xff)
        txd_data[41] = (int(entryVar['PID_Ctrl1[4]'].get()) & 0xff)
        txd_data[42] = (int(entryVar['PID_Ctrl1[5]'].get()) & 0xff)
        txd_data[43] = (int(entryVar['PID_Ctrl1[6]'].get()) & 0xff)
        txd_data[44] = (int(entryVar['PID_Ctrl1[7]'].get()) & 0xff)
        txd_data[45] = (int(entryVar['PID_Ctrl1[8]'].get()) & 0xff)
    
        txd_data_t = txd_data[0:(len(txd_data) - offset)]       
        #CRC_Calculate_txd_data(txd_data_t, offset)
            
        CRC = 0xFFFF
        #for i in range(0, CPBYTELength + 4 - 2 - offset):
        for i in range(0, len(txd_data_t) - 2):
            CRC ^= txd_data_t[i]
            for j in range(0, 8):
                if (CRC & 0x01):
                    CRC= (CRC >> 1) ^ 0x8408
                else:
                    CRC >>= 0x01
        CRC_txd = (~CRC) & 0xffff   

        txd_data_t[len(txd_data_t) - 2] = (CRC_txd >> 8) & 0xff #高位
        txd_data_t[len(txd_data_t) - 1] = CRC_txd & 0xff        #低位

        ser.reset_input_buffer()
        try:
            #txd_data_t = txd_data[0:(len(txd_data) - offset)]
            ser.write(txd_data_t)
            #print("APP txd:")
            #print(txd_data_t)
            if txd_data_t[1] == 4:
                
                print(len(txd_data_t))                
        except:
            pass
            print('quit txdTask2 \r\n')
            #ser.close()
            #sys.exit()
        #if sendRequest == 0:
            #recoverCnt = 10
        #sendRequest = 0
        #checkButton["Edit"].deselect()
        time.sleep(0.5)    #(0.25)
        sendRequest = 1
        
        '''
        if txd_data[1] == 4:
            time.sleep(0.25)
            checkButton["Edit"].deselect() 
        '''
  
#验证  no use
def verificationTask():
    global LoadCPEnable
    global loadModel 
    ok = 0
    
    #loadModel = (int(entryVar['Software Model'].get()) & 0xff)
    try:
        tmp = entryVar['CPFile'].get() + '.xlsx'    #从Entry控件中获取文本，后面紧跟着 .xlsx 的文件名 
        print(tmp)  
        wb_cp = openpyxl.load_workbook(tmp)    #该函数会读取文件内容，并返回一个 Workbook 对象，你可以通过这个对象来访问和操作工作簿中的工作表、单元格等数据。
        print(wb_cp)
        sheet_cp = wb_cp.get_sheet_by_name('Control Flow Parameter')    #访问工作表
        print(sheet_cp)
        for i in range(0, CPLength - cplength_offset): 
            tmp = sheet_cp.cell(row = i + 2, column = 2 + loadModel - 1).value    #访问工作表中特定单元格的值
            print(entryVar[RxdTabs[i]].get())           
            #entryVar[RxdTabs[i]].set(tmp)
            if (entryVar[RxdTabs[i]].get()) != tmp:        
                ok = 1
                print("verification failed?")   #验证
                break

        if ok == 0:
            print("verification success")
            messagebox.showinfo('Pass',"Verification Success.")   #用于显示一个包含信息消息和标题的模态对话框
        else:
            print("verification failed") 
            messagebox.showinfo('Failed',"Verification Failed.")
    except:
        print ("Error: can\'t find folder or file can't be uploaded")
        messagebox.showinfo('Error',"can\'t find folder or file can't be uploaded.")
        time.sleep(5)
    
def t05sTask():
    global lostCnt
    global receivingnotok
    global recoverCnt
    global sendRequest
    global verificationEnable
    global sendRequestCnt
    lostCnt = 0
    sendRequestCnt = 0
    while True:
        #if receivingnotok:
            #receivingnotok = 0
            #if lostCnt < 6:
                #lostCnt = lostCnt + 1
                #if lostCnt >= 3:
                    #lostCnt = 0
                    #ser.reset_input_buffer()
                    
        #time.sleep(0.5)
        if verificationEnable > 1:
            verificationEnable -= 1
        
        sendRequestCnt = sendRequestCnt + 1
        if sendRequestCnt >= 5:
            sendRequestCnt = 0
            #sendRequest = 1
        
        #if recoverCnt:
            #recoverCnt = recoverCnt - 1
            #if recoverCnt == 0:
                #checkButton["Edit"].deselect()
        
        if portOK:
            lostCnt = lostCnt + 1
            if lostCnt >= 30:
                lostCnt = 0
                ser.reset_input_buffer()
    
        time.sleep(0.05)        
        #print("lostCnt = ", lostCnt)

#*********main*********
portOK = 0
top = tkinter.Tk()
top.title("Intellenget Oven")

_thread.start_new_thread(topTask, ())
_thread.start_new_thread(PortTask, ())
_thread.start_new_thread(t05sTask, ())

top.mainloop()



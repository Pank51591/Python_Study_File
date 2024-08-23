# Author: Peng Yuan
# V05 20200902 --- Add MaxAugerCurrent and MinAugerCurrent that will be received through RS232, replace the receiving of the waterDspAux, and other 3 
#                  spareSwitch.
# V06 20221-13 --- Added spRunOnSb,spStartupDly,icePerRes,DlyWUsageSb and iceCapacity.
# V07 20230627 --- Added PhD model


import serial
import sys
import time
import serial.tools.list_ports
import _thread
import os
import tkinter
import openpyxl
import datetime
from tkinter import *
#import pandas
from tkinter import messagebox
from numbers import Number
from decimal import Decimal
import ctypes
from ctypes import c_int8
from ctypes import c_int16
from distutils.core import setup
#import tkfont
#import py2exe

# start address, length
#SaveAddress = 69
CPBYTELength = (101 + 7 + 5 + 5 + 1) # change  （字节个数）
#ENTRY
CPLength = (76 + 1 + 4 + 1 + 1) # change  （参数个数）
PPLength = 8
StatusLength = (21 + 2)
OutstatusLength = (9)
#MAX_ENTRY = CPLength + PPLength + StatusLength
#SaveLength = PPLength + StatusLength + OutstatusLength
MAX_RECV_BYTE = (170 + 7 + 5 + 2 + 5 + 1) # change  （接收的字节数）

TEST_DATA = [000000,000000,000000,0x000100,0x000200,0x000400,0x000800,0x001000,0x002000,0x004000,0x008000,
             0x800000,0x400000,0x200000,0x100000,0x080000,0x040000,0x010000,0x020000,
             0x000001,0x000002,0x000004,0x000008,0x000010,0x000020,0x000040,0x000080,000000,000000,000000]

#rxd
rxd_data = [0] * MAX_RECV_BYTE

#txd 
txd_data = [0] * (CPBYTELength + 4) #105
entry = {}  
entryVar ={}
    
checkVar = {}
checkButton = {}

verif_ok = 0
keyNum = 0
keyDelay = 0
sendEnable = 0
commCnt = 0
lostCnt = 0
Reget = 0
rowIndex = 2
startTest = 0
PersistentParameterReset = 0
saveEnable = 0
saveCPEnable = 0
LoadCPEnable = 0
receivingnotok = 0
lost = 0
recoverCnt = 0
verificationEnable = 0
CRC = 0
CRC_rxd = 0
sendRequest = 0
editOnce = 0
sendRequestCnt = 0
offset = 0
startup = 0
cplength_offset = 0

OUTPUT_STATUS = ('Compressor','Fan','Auger','FailSafe','Drain','WaterDispenser','IceDispenser','ExtraOutput','UVLamp')

RxdTabs = (
         'augerCurMax','augerCurMin','augerCurMinEn','augerRunOnSb','augerTDRunOnSb',\
         'binSignalLowHigh','callWaterToutSw','cleaningBits','cleanTimeMb','compServiceTHw',\
         'compStartupDly','dlyAfterPourSb','dlyBeforeEmptySb','dlyBeforeRinseSb','dip',\
         'dispIMDlyDur1Tw','dispIMDlyDur2Tw','dispIMDlyDur3Tw','dispIMDlyDur4Tw','callW.dlyDurMw',\
         'callW.dlyMonHb','callW.dis','hAmps.dlyDurMw','hAmps.dlyMonHb','hAmps.dis',\
         'hPres.dlyDurMw','hPres.dlyMonHb','hPres.dis','lowW.dlyDurMw','lowW.dlyMonHb',\
         'lowW.dis','wLeak.dlyDurMw','wLeak.dlyMonHb','wLeak.dis','drainToutSb',\
         'fanRunOnMb','fanStartupDly','gearSensingDly','gearStartupDly','have',\
         'hideUItext','imStartupDly','lowWDlySb','maxOffTMw','model',\
         'modelInd','pDebounceBits','selfFlushBits1','selfFlushBits2','selfFlushDur1Sw',\
         'selfFlushDur2Sw','selfFlushFreq1Sw','selfFlushFreq2Sw','selfFlushRepeat1','selfFlushRepeat2',\
         'sleepDelayMb','lightUIDurSleep','tdlyDrainToutSb','tDlyFlushDip','tDlyFlushDurSb',\
         'tDlyFlushEn','tDlyFlushFreqMb','tDlyFlushRep','tDlyLongMb','tDlyShortMb',\
         'FailsafeWhenDisp','waterManage','wMDrainRetryMw','waterFillToutSw','modbusAddress',\
         'noWaterADLo','noWaterADHi','noWaterADDrip','noWaterADLeak','shuttleFlushEn',\
         'callWaterToutOnStartSw','sn','spRunOnSb','spStartupDly','icePerRes','DlyWUsageSb','PhDModel','OverCurrentCnt','compCycCountdw',\
         'compOnTimeHw','iDispCycCntdw','iceDispTMdw','powerOnTMdw','sixMonthMdw',\
         'wDispCycCntdw','wDispTimeMdw','augerCurrent','lowWaterSen','HiwWaterSen',\
         'dripWaterSen','leakWaterSen','binSwitch','tdsSwitch','hiPresSwitch',\
         'cleanSwitch','iceDispAux','maxAugerCurrent','minAugerCurrent',\
         'dipSwitch','outState','errState','myMode','version',\
         'LO_CHN_AD','HI_CHN_AD','DRIP_CHN_AD','LEAK_CHN_AD',\
         'tmSinceLstFls','iceCapacity',\
         'T_AU','T_AV','T_AW','T_AX',\
         'T_U1','T_V1','T_W1','T_X1','T_Y1',\
         'T_Z1','T_AA1','T_AB1','T_AC1','T_AD1',\
         'T_AE1','T_AF1','T_AG1','T_AH1','T_AI1',\
         'T_AJ1','T_AK1','T_AL1','T_AM1','T_AN1',\
         'T_AE2','T_AF2','T_AG2','T_AH2','T_AI2',\
         'T_AJ2','T_AK2','T_AL2','T_AM2','T_AN2',\
         'T_AO1','T_AP1','T_AQ1','T_AR1','T_AS1',\
         'T_AT1_','T_AU1','T_AV1','T_AW1','T_AX1',\
         'T_AO2','T_AP2','T_AQ2','T_AR2','T_AS2',\
         'T_AT2_','T_AU2','T_AV2','T_AW2','T_AX2',         )         
         
SaveTabs = ('   Date      Time',\
         'compCycCountdw',\
         'compOnTimeHw','iDispCycCntdw','iceDispTMdw','powerOnTMdw','sixMonthMdw',\
         'wDispCycCntdw','wDispTimeMdw','augerCurrent','lowWaterSen','HiwWaterSen',\
         'dripWaterSen','leakWaterSen','binSwitch','tdsSwitch','hiPresSwitch',\
         'cleanSwitch','iceDispAux','maxAugerCurrent','minAugerCurrent',\
         'dipSwitch','outState','errState','myMode','version',\
         'LO_CHN_AD','HI_CHN_AD','DRIP_CHN_AD','LEAK_CHN_AD','tmSinceLstFls','iceCapacity',\
         'Compressor','Fan','Auger','FailSafe','Drain','WaterDispenser','IceDispenser','ExtraOutput','UVLamp') 

Baudlist = ('300','9600')

def CRC_Calculate_txd_data(*data, offset): #no use
    global CRC_txd
    global txd_data

    CRC = 0xFFFF
    for i in range(0, CPBYTELength + 4 - 2 - offset):
        #CRC ^= data[i]
        for t in data: 
            CRC ^= t 
        for j in range(0, 8):
            if (CRC & 0x01):
                CRC= (CRC >> 1) ^ 0x8408
            else:
                CRC >>= 0x01
    CRC_txd = (~CRC) & 0xffff


#接收数据CRC校验
def CRC_Calculate_rxd_data(bytes):
    global ss
    global CRC_rxd

    CRC = 0xFFFF
    #for i in range(0, MAX_RECV_BYTE - 2):
    for i in range(0, bytes - 2):
        CRC ^= (ss[i])
        for j in range(0, 8):
            if (CRC & 0x01):
                CRC= (CRC >> 1) ^ 0x8408
            else:
                CRC >>= 0x01
    CRC_rxd = (~CRC) & 0xffff
    

def MlKeyTask():
    global keyNum
    global keyDelay
    
    keyNum = 7
    keyDelay = 1
    
def MrKeyTask():
    global keyNum
    global keyDelay
    
    keyNum = 6
    keyDelay = 1
    
def MprKeyTask():
    global keyNum
    global keyDelay
    
    keyNum = 5
    keyDelay = 1
    #messagebox.showinfo( "Off", keyNum)

def MinusKeyTask():
    global keyNum
    global keyDelay
    
    keyNum = 4
    keyDelay = 1
    #messagebox.showinfo( "Off", keyNum)   
    
def PlusKeyTask():
    global keyNum
    global keyDelay
    
    keyNum = 3
    keyDelay = 1
    #messagebox.showinfo( "Off", keyNum)


def OffKeyTask():
    global keyNum
    global keyDelay
    
    keyNum = 2
    keyDelay = 1
    #messagebox.showinfo( "Off", keyNum)

def MsKeyTask():
    global keyNum
    global keyDelay
    
    keyNum = 1
    keyDelay = 1
    #messagebox.showinfo( "Ms", keyNum + keyDelay)
  
def StartTestTask():
    global startTest
    
    startTest = 1
    #messagebox.showinfo( "Ms", keyNum + keyDelay)

#设置按钮的函数
def putButton(row_s, col_s, name, task):
    button ={}
    
    button[name] = tkinter.Button(top, text = name, command = task, width = 14)  #按钮控件，task 按键按下之后会调用的函数
    button[name].grid(row = row_s, column = col_s)

#设置标签
def putLable(row_s, col_s, name):
    lable = {}
    lable[name] = Label(top, text = name, width = 14, wraplength = 0)
    lable[name].grid(row = row_s, column = col_s)

#
def putEntry(row_s, col_s, name, value, sta):
    global entryVar
    global entry

    entryVar[name] = StringVar()
    entry[name] = Entry(top, bd = 4, textvariable = entryVar[name], width = 14, state = sta)#
    entry[name].grid(row = row_s, column = col_s)
    entryVar[name].set(value)
    entry[name].select_range(0, 256 )

#设置选择框
def putCheckbutton(row_s, col_s, name):
    global checkButton
    global checkVar
    
    checkVar[name] = IntVar()
    checkButton[name] = Checkbutton(top, text = name, variable = checkVar[name], \
        onvalue=1, offvalue=0)
    checkButton[name].var = checkVar[name]
    checkButton[name].grid(row = row_s, column = col_s)

def SendTask():
    global sendEnable
    global disable15MSend
    
    sendEnable = 1
    disable15MSend = 1

#
def verificationTask():
    global LoadCPEnable
    global loadModel 
    ok = 0
    
    #loadModel = (int(entryVar['Software Model'].get()) & 0xff)
    if loadModel >= 1 and  loadModel <= 15:
        try:
            tmp = entryVar['CPFile'].get() + '.xlsx'    #从Entry控件中获取文本，后面紧跟着 .xlsx 的文件名 
            print(tmp)  
            wb_cp = openpyxl.load_workbook(tmp)  #该函数会读取文件内容，并返回一个 Workbook 对象，你可以通过这个对象来访问和操作工作簿中的工作表、单元格等数据。
            print(wb_cp)
            sheet_cp = wb_cp.get_sheet_by_name('Control Flow Parameter')  #访问工作表
            print(sheet_cp)
            for i in range(0, CPLength - cplength_offset): 
                tmp = sheet_cp.cell(row = i + 2, column = 2 + loadModel - 1).value
                print(entryVar[RxdTabs[i]].get())           
                #entryVar[RxdTabs[i]].set(tmp)
                if (entryVar[RxdTabs[i]].get()) != tmp:        
                    ok = 1
                    print("verification failed?")   #验证
                    break

            if ok == 0:
                print("verification success")
                messagebox.showinfo('Pass',"Verification Success.")   #用于显示一个包含信息消息和标题的模态对话框
            else:
                print("verification failed") 
                messagebox.showinfo('Failed',"Verification Failed.")
        except:
            print ("Error: can\'t find folder or file can't be uploaded")
            messagebox.showinfo('Error',"can\'t find folder or file can't be uploaded.")
            time.sleep(5)
    else:
        messagebox.showinfo('Error',"Software Model should be 1 ~ 15!")
        time.sleep(5)

#   
def ProgrammingMachineTask():
    global sendEnable
    global disable15MSend
    global LoadCPEnable
    global verificationEnable
    global editOnce
    
    checkButton["Edit"].select()
    while checkButton["Edit"].var.get() == 0:
        time.sleep(1)
    LoadCPEnable = 2
    LoadCPTask()
    print("verificationEnable:", verificationEnable)    

#     
def ProgrammingTask():
    global sendEnable
    global disable15MSend
    global LoadCPEnable
    global verificationEnable
    global editOnce
    
    #checkButton["Disable 15M"].deselect()
    checkButton["Edit"].select()
    while checkButton["Edit"].var.get() == 0: # or checkButton["Disable 15M"].var.get():
        time.sleep(1)
    LoadCPEnable = 3
    LoadCPTask()
    #sendEnable = 0
    #time.sleep(1)
    #ResetPersistentParameterTask()
    #disable15MSend = 1 

    #sendEnable = 2
    #time.sleep(1)
    #checkButton["Edit"].deselect()
    #verificationEnable = 1
    print("verificationEnable:", verificationEnable)

#    
def autoLengthTask():
    startup = 0
    
def defaultTask():
    global sendEnable
    global disable15MSend
    global LoadCPEnable
    global verificationEnable
    global sendRequest
    global sendRequestCnt 
    
    #checkButton["Disable 15M"].deselect()
    #checkButton["Edit"].select()
    #while checkButton["Edit"].var.get() == 0:# or checkButton["Disable 15M"].var.get():
    #    time.sleep(1)
    
    #time.sleep(1)
    #ResetPersistentParameterTask()
    #disable15MSend = 1   
    #time.sleep(1)
    #sendRequest = 0
    #sendRequestCnt = 0    
    sendEnable = 3
    
    
def ResetPersistentParameterTask():
    global PersistentParameterReset
    
    PersistentParameterReset = 1
    
def ClearSaveBufferTask():
    rowIndex = 2 
    
def putBlankText(row_s):
    text = Text(top, width = 17, height = 1, bg = 'SystemButtonFace', relief = FLAT)
    text.insert(INSERT, "")
    #text.insert(END, "")
    text.grid(row = row_s, column = 0)
    
    #defaultbg = top.cget('bg')
    #print(defaultbg)
    
def SaveCPEnableTask():
    global saveCPEnable
    saveCPEnable = 1
    
def SaveCPTask():
    global saveCPEnable
    
    while True:
        if saveCPEnable:
            model = (int(entryVar['model'].get()) & 0xff)
            if model >= 1 and model <= 15:
                saveCPEnable = 0
                sheet_cp.cell(row = 1, column = 1).value = "model No."  
                for i in range(0, 15):
                    sheet_cp.cell(row = 1, column = i + 2).value = i + 1  
                for i in range(0, CPLength): 
                    # items
                    sheet_cp.cell(row = i + 2, column = 1).value = RxdTabs[i]
                    # parameters by model
                    sheet_cp.cell(row = i + 2, column = 2 + model - 1).value = entryVar[RxdTabs[i]].get()     
                fileName_cp = entry["CPFile"].get() + '.xlsx'
                try:
                    wb_cp.save(fileName_cp)
                    print("saved Control Flow Parameter")
                    messagebox.showinfo('Saved',"The Control Flow Parameter is saved.") 
                    #time.sleep(5)
                except:
                    print ("Error: can\'t find folder or file is open and can't be saved")
                    messagebox.showinfo('Error',"can\'t find folder or file is open and can't be saved.")
                    time.sleep(5)
            else:
                messagebox.showinfo('Error',"model should be 1 ~ 15!")
                time.sleep(5)                  
        time.sleep(1)

def LoadCPEnableTask(): 
    global LoadCPEnable
    
    checkButton["Edit"].select()
    LoadCPEnable = 1
    LoadCPTask()
 
 #15个机型的具体编号
MODEL_1 = ('00938605','00966887','00966895','00958033','99999991') 
MODEL_2 = ('00965681','00965699','00965707','00969048','99999992','01102227','01055243','01133685')
MODEL_3 = ('00976795','00976803','01006428','01016781','99999993')
MODEL_4 = ('00976811','00976829','01006410','99999994')
MODEL_5 = ('01016724','01016732','01026582','99999995')
MODEL_6 = ('01074541','01074558','01074608','01074616','01113380','01074566','01074574','01074624','01074632','01113398','99999996','01117811')
MODEL_7 = ('00997999','01006741','01038264','01038272','01002419','01036110','01036144','01064708','01070176','01070184','99999997')
MODEL_8 = ('99999998','01095405','01095413','01095421','01095439','01096353','01133842','01139245','01139252','01140078','01140060','01111657','01070150','01070168','01070192','01070200')
MODEL_9 = ('01049436','01051978','99999999','01210368')
MODEL_10 = ('01049444','01051986','99999910','01102235','01055342','01133677')
MODEL_11 = ('01074582','01074590','99999911','01117829')
MODEL_12 = ('01125129','01125137','99999912','01165679')
MODEL_13 = ('01200864','01222827','99999913')
MODEL_14 = ('99999914','99999914')
MODEL_15 = ('99999915','99999915')

#验证
def verification():
    global LoadCPEnable
    global loadModel 
    global verif_ok
    
    verif_ok = 0
    if loadModel >= 1 and  loadModel <= 15:
        try:
            tmp = entryVar['CPFile'].get() + '.xlsx'   
            print(tmp)  
            wb_cp = openpyxl.load_workbook(tmp)
            print(wb_cp)
            sheet_cp = wb_cp.get_sheet_by_name('Control Flow Parameter')
            print(sheet_cp)
            for i in range(0, CPLength - cplength_offset): 
                tmp = sheet_cp.cell(row = i + 2, column = 2 + loadModel - 1).value
                           
                #entryVar[RxdTabs[i]].set(tmp)
                if (entryVar[RxdTabs[i]].get()) != tmp:        
                    verif_ok = 1
                    print("verification failed?") 
                    break
            if verif_ok == 0:
                print("verification success_2")
            else:
                print("verification failed_2") 
        except:
            pass
            #print ("Error: can\'t find folder or file can't be uploaded")
           # messagebox.showinfo('Error',"can\'t find folder or file can't be uploaded.")
            #time.sleep(5)
    else:
        messagebox.showinfo('Error',"Software Model should be 1 ~ 15!")
        time.sleep(5)
        
def LoadCPTask():
    global verif_ok
    global LoadCPEnable
    global verificationEnable
    global sendEnable
    global loadModel
    
    #while True:
    if LoadCPEnable:   
        if LoadCPEnable == 3:
            loadModel = (int(entryVar['Software Model'].get()) & 0xff)
        elif LoadCPEnable == 2:
            loadModelBuf = entryVar['Machine Model'].get()
            loadModel = 0
            tmp = 'Model2SoftModel.xlsx'   
            print(tmp)  
            try:
                wb_model = openpyxl.load_workbook(tmp)
                print(wb_model)
                sheet_model = wb_model.get_sheet_by_name('Models')
                print(sheet_model)
            except:
                messagebox.showinfo('Error',"No Model2SoftModel.xlsx!")
            find = 0
            for k in range(0, 15):
                 
                tmp2 = sheet_model.cell(row = 1, column = k + 2).value
                print(tmp2)
                print(len(tmp2))
               
                model = ''
                cnt = 0
                for j in range(0, len(tmp2)):
                    if tmp2[j] != '\n':
                        model += tmp2[j]
                        cnt += 1
                        if cnt == 8:
                            cnt = 0
                            if loadModelBuf == model:
                                print(loadModelBuf)
                                find = 1
                                loadModel = k + 1
                                break;
                            model = ''
                if find:
                    break                            
                
        if loadModel >= 1 and  loadModel <= 15:
            try:
                tmp = entryVar['CPFile'].get() + '.xlsx'   
                print(tmp)  
                wb_cp = openpyxl.load_workbook(tmp)
                print(wb_cp)
                sheet_cp = wb_cp.get_sheet_by_name('Control Flow Parameter')
                print(sheet_cp)
                for i in range(0, CPLength - cplength_offset): 
                    tmp = sheet_cp.cell(row = i + 2, column = 2 + loadModel - 1).value
                    tmp1 = sheet_cp.cell(row = 77 + 2, column = 2 + loadModel - 1).value
                    print(tmp1)
                    if tmp != None:
                        entryVar[RxdTabs[i]].set(tmp)
                    else:
                        messagebox.showinfo('Error',"Invalid Value in *.xlsx.")
                        LoadCPEnable = 0
                        #if sendEnable == 2:
                            #sendEnable = 0
                            #verificationEnable = 0
                        break
                if LoadCPEnable == 2 or LoadCPEnable == 3:
                    sendEnable = 2
                    verificationEnable = 20
                print("Upload Control Flow Parameter")
                if LoadCPEnable == 1:
                    messagebox.showinfo('Upload',"The Control Flow Parameter is upload.") 
                #time.sleep(5)
            except:
                print ("Error: can\'t find folder or file can't be uploaded")
                messagebox.showinfo('Error',"can\'t find folder or file can't be uploaded.")
                time.sleep(5)
        else:
            if LoadCPEnable == 2:
                messagebox.showinfo('Error',"No this Machine Model Numbers!")
            else :
                messagebox.showinfo('Error',"Software Model should be 1 ~ 15!")
        LoadCPEnable = 0
            #time.sleep(5)
        #verification()
        #if verif_ok == 0:
            #LoadCPEnable = 0
        #else:
            #verif_ok = 1 #test
        #time.sleep(0.1)  

   
def topTask():
    #putCheckbutton(0, 1, "Port Open")
    
    putEntry(0, 3, "SaveFile", "SaveData", 'normal')
    #putCheckbutton(0, 4, "Save")
    putButton(0, 4, "Save", SaveEnableTask)
    
    #putLable(0, 5, 'Send Cycle [s]')
    #putEntry(0, 6, "Send Cycle", 0.5, 'normal')
    #putCheckbutton(0, 7, "Auto Send")
    putButton(0, 5, "Send Once", SendTask)
    putCheckbutton(0, 6, "Edit")
    putButton(0, 7, "Reset Persistent", ResetPersistentParameterTask)
    #putButton(0, 9, "Start Test", StartTestTask)
    #putLable(0, 10, 'Test Interval[0.1s]')
    #putEntry(0, 11, "Send Speed", 10, 'normal')    
    putEntry(0, 8, "CPFile", "Control Parameter", 'normal')
    putButton(0, 9, "Save CP", SaveCPEnableTask) 
    putButton(0, 10, "Load CP", LoadCPEnableTask) 
    putCheckbutton(0, 11, "Disable 15M")
    putCheckbutton(1, 0, "Receiving")
    #putButton(1, 1, "Auto Length", autoLengthTask)
    putButton(1, 5, "Default CP", defaultTask)
    putLable(1, 6, 'Software Model')
    putEntry(1, 7, 'Software Model', 1, 'normal')   
    putButton(1, 8, "Programming_s", ProgrammingTask)
    
    putLable(1, 9, 'Machine Model')
    putEntry(1, 10, 'Machine Model', '00938605', 'normal')    
    putButton(1, 11, "Programming_m", ProgrammingMachineTask)
    
    #putBlankText(1)
    
    #putLable(2, 0, 'Max Send Byte')
    #putEntry(2, 1, "Max Send Byte", 50, 'normal')
    
    #putButton(2, 3, "Clear Save Buffer", ClearSaveBufferTask)
         
    #i = 0
    #j = 0
    #for lTab in RxdTabs:
       #putLable(10 + j * 2, i, lTab)
       #putEntry(11 + j * 2, i, lTab, 0, 'readonly')  
       #i = zi + 1
       ## change row at which column
       #if i == 10:
           #i = 0
           #j = j + 1       
           
    #i = 0
    #j = 0
    #num = 0
    #section = 0
    #putLable(9, 5, "******************")
    #putLable(9, 7, "******************")
    #putLable(9, 6,  "Control para")    
    #for lTab in RxdTabs:
        #if section == 0:
            #putLable(10 + j, i * 2, lTab)
            #if num < CPLength:
                #putEntry(10 + j, i * 2 +  1, lTab, 0, 'norm')
            #else:
                #putEntry(10 + j, i * 2 +  1, lTab, 0, 'readonly')
            #num = num + 1
            #i = i + 1
            ## change row at which column
            #if i == 6:
                #i =  0
                #j = j + 1
            #if num == CPLength or num == CPLength + PPLength or num == CPLength + PPLength + StatusLength: 
                #i = 0   
                #j = j + 1
                #putBlankText(22)
                #putLable(23, 5, "******************")
                #putLable(23, 7, "******************")
                #putLable(23, 6,  "Persistent para")
                #putBlankText(26)
                #putLable(27, 5, "******************")
                #putLable(27, 7, "******************")
                #putLable(27, 6,  "Running Status")
                #j = j + 2
            #if num == MAX_ENTRY:
                #break
    # CP            
    i = 0
    j = 0
    num = 0
    putLable(9, 5, "******************")
    putLable(9, 7, "******************")
    putLable(9, 6,  "Control para")                    
    for lTab in RxdTabs:
        putLable(10 + j, i * 2, lTab)
        putEntry(10 + j, i * 2 +  1, lTab, 0, 'norm')
        num = num + 1
        j = j + 1
        # change row at which column
        if j == 13:
            j =  0
            i = i + 1
        if num == CPLength:
            break
    # PP    
    i = 0
    j = 0
    num = 0
    #putLable(22, 5, "")
    putLable(23, 5, "******************")
    putLable(23, 7, "******************")
    putLable(23, 6,  "Persistent para")                    
    for lTab in RxdTabs:
        num = num + 1
        if num > CPLength:
            putLable(24 + j, i * 2, lTab)
            putEntry(24 + j, i * 2 +  1, lTab, 0, 'readonly')
            
            i = i + 1
            # change row at which column
            if i == 6:
                i =  0
                j = j + 1
        if num == CPLength + PPLength:
            break   
        
    # RUN Status    
    i = 0
    j = 0
    num = 0
    #putLable(26, 5, "")
    putLable(27, 5, "******************")
    putLable(27, 7, "******************")
    putLable(27, 6,  "Runing Status")                    
    for lTab in RxdTabs:
        num = num + 1
        if num > CPLength + PPLength:
            putLable(28 + j, i * 2, lTab)
            putEntry(28 + j, i * 2 +  1, lTab, 0, 'readonly')
            
            j = j + 1
            # change row at which column
            if j == 4:
                j =  0
                i = i + 1
        #if num == CPLength + PPLength + StatusLength:
       #     j = 4
        #    i = 0
        if num == CPLength + PPLength + StatusLength:
            break   
    #version    
    
    # OUT Status 
    #putLable(32, 5, "")
    putLable(33, 5, "******************")
    putLable(33, 7, "******************")
    putLable(33, 6,  "Output Status")    
    j = 0
    i = 0
    for lTab in OUTPUT_STATUS:
        putLable(34 + j, i * 2, lTab)         
        putEntry(34 + j, i * 2 +  1, lTab, 0, 'readonly')
        i = i + 1
        if i == 6:
            i = 0
            j = j + 1

#            
def excelTask():
    global sheet
    global wb
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'SerialForm'
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 15
    sheet.column_dimensions['H'].width = 15 
    sheet.column_dimensions['I'].width = 15
    sheet.column_dimensions['J'].width = 15
    sheet.column_dimensions['K'].width = 15
    sheet.column_dimensions['L'].width = 15
    sheet.column_dimensions['M'].width = 15
    sheet.column_dimensions['N'].width = 15
    sheet.column_dimensions['O'].width = 15
    sheet.column_dimensions['P'].width = 15   
    sheet.column_dimensions['Q'].width = 15
    sheet.column_dimensions['R'].width = 15
    sheet.column_dimensions['S'].width = 15
    sheet.column_dimensions['T'].width = 15    
    sheet.column_dimensions['U'].width = 15
    sheet.column_dimensions['V'].width = 15 
    sheet.column_dimensions['W'].width = 15
    sheet.column_dimensions['X'].width = 15  
    sheet.column_dimensions['Y'].width = 15
    sheet.column_dimensions['Z'].width = 15
    sheet.column_dimensions['AA'].width = 15
    sheet.column_dimensions['AB'].width = 15
    sheet.column_dimensions['AC'].width = 15
    sheet.column_dimensions['AD'].width = 15
    sheet.column_dimensions['AE'].width = 15
    sheet.column_dimensions['AF'].width = 15
    sheet.column_dimensions['AG'].width = 15
    sheet.column_dimensions['AH'].width = 15
    sheet.column_dimensions['AI'].width = 15
    sheet.column_dimensions['AJ'].width = 15
    sheet.column_dimensions['AK'].width = 15
    sheet.column_dimensions['AL'].width = 15    
    
    #sheet.freeze_panes = 'A2'
    sheet.freeze_panes = 'B2'
    
    global sheet_cp
    global wb_cp
    
    wb_cp = openpyxl.Workbook()
    sheet_cp = wb_cp.active
    sheet_cp.title = 'Control Flow Parameter'
    sheet_cp.column_dimensions['A'].width = 20
    sheet_cp.column_dimensions['B'].width = 15
    #sheet_cp.column_dimensions['C'].width = 15
    #sheet_cp.column_dimensions['D'].width = 15
    #sheet_cp.column_dimensions['E'].width = 15
    #sheet_cp.column_dimensions['F'].width = 15
    #sheet_cp.column_dimensions['G'].width = 15
    #sheet_cp.column_dimensions['H'].width = 15 
    #sheet_cp.column_dimensions['I'].width = 15
    #sheet_cp.column_dimensions['J'].width = 15
    #sheet_cp.column_dimensions['K'].width = 15
    #sheet_cp.column_dimensions['L'].width = 15
    #sheet_cp.column_dimensions['M'].width = 15
    #sheet_cp.column_dimensions['N'].width = 15
    #sheet_cp.column_dimensions['O'].width = 15
    #sheet_cp.column_dimensions['P'].width = 15   
    #sheet_cp.column_dimensions['Q'].width = 15
    #sheet_cp.column_dimensions['R'].width = 15
    #sheet_cp.column_dimensions['S'].width = 15
    #sheet_cp.column_dimensions['T'].width = 15    
    #sheet_cp.column_dimensions['U'].width = 15
    #sheet_cp.column_dimensions['V'].width = 15 
    #sheet_cp.column_dimensions['W'].width = 15
    #sheet_cp.column_dimensions['X'].width = 15  
    #sheet_cp.column_dimensions['Y'].width = 15
    #sheet_cp.column_dimensions['Z'].width = 15
    #sheet_cp.column_dimensions['AA'].width = 15
    #sheet_cp.column_dimensions['AB'].width = 15
    #sheet_cp.column_dimensions['AC'].width = 15
    #sheet_cp.column_dimensions['AD'].width = 15
    #sheet_cp.column_dimensions['AE'].width = 15
    #sheet_cp.column_dimensions['AF'].width = 15
    #sheet_cp.column_dimensions['AG'].width = 15
    #sheet_cp.column_dimensions['AH'].width = 15
    #sheet_cp.column_dimensions['AI'].width = 15
    #sheet_cp.column_dimensions['AJ'].width = 15
    #sheet_cp.column_dimensions['AK'].width = 15
    #sheet_cp.column_dimensions['AL'].width = 15    
    sheet_cp.freeze_panes = 'B2'    

#rxd_data = [90,18,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,26,16]


        
def SaveTask():
    global csvInit
    global saveEnable
    #try:
        #c = checkButton["Save"].var.get()
    #except:
        #print('quit rxdTask4 \r\n')
        #sys.exit()
    #saveEnable = 1  
    while True:
        if saveEnable: 
            saveEnable = 0
            fileName = entry["SaveFile"].get() + '.xlsx'
            #try:
                #wb.save(fileName)
                #print("saved2")
                #messagebox.showinfo('Saved',"The file is saved.") 
                #time.sleep(5)
            #except:
                #print ("Error: can\'t find folder or file is open and can't be saved")
                #messagebox.showinfo('Error',"can\'t find folder or file is open and can't be saved.")
                #time.sleep(5)
                
            try: 
                #os.rename(fileName, 'tempfile.xls')
                #os.rename('tempfile.xls', fileName)
                wb.save(fileName)
                print("saved2")
                messagebox.showinfo('Saved',"The file is saved.") 
                #time.sleep(5)            
            except: # OSError:
                print ("Error: The file is still open and can't be saved. Please close it first!")
                messagebox.showinfo('Error',"The file is still open and can't be saved, please close it first!")
                time.sleep(5)            
        time.sleep(1)         

def SaveEnableTask():
    global saveEnable
    saveEnable = 1
    #SaveTask()
    
def txdTask():
    global sendEnable
    global disable15MSend
    global PersistentParameterReset
    global keyNum
    global keyDelay
    global ser
    global portOK
    global rowIndex
    global sheet_cp    
    global startTest
    global recoverCnt
    global CRC_txd
    global sendRequest
    global sendRequestCnt
    global offset
    global startup
    
    startup = 0
    disable15MSend = 0
    cnt = 0
    time.sleep(0.5)
    while portOK:
        delay = 0
        first = 0
        
        #tp = [0] * 6
        
        #tp[0] = 'I'#(entryVar['sn'].set(ss2[i + r_offset]))
        #tp[1] = '0'#(entryVar['sn'].set(ss2[i + r_offset]))
        #tp[2] = '0'#(entryVar['sn'].set(ss2[i + r_offset]))
        #tp[3] = '0'#(entryVar['sn'].set(ss2[i + r_offset]))
        #tp[4] = '0'#(entryVar['sn'].set(ss2[i + r_offset]))
        #tp[5] = '1'#(entryVar['sn'].set(ss2[i + r_offset]))
        ##tp[6] = '/0'
        
        #strtp = ''.join(str(e) for e in tp)
        
        #entryVar['sn'].set(strtp)        
        
        c = startTest
        try:
            #if c:
            maxByte = CPBYTELength + 4 # no use
            #else:
            #     maxByte = 4
        except:
            pass
            messagebox.showinfo('error', 'not a int number!')
            time.sleep(5) 
            
        txd_data[0] = 0x5a

        
        if c:
            try:
                d = int(entryVar['Send Speed'].get())
            except:
                pass
                messagebox.showinfo('error', 'not a number!')
                time.sleep(5)
        #if c:
            #if first:
                #first = 1
                #sendEnable = 1
            #else:
                #time.sleep(d)
                #sendEnable = 1
        #else:
            #first = 0
           
        #print(startTest)
        #print(txd_data[2])
        if  sendRequest or sendEnable or disable15MSend or PersistentParameterReset:  
        #if  sendRequest:
            #if startup == 0:
                #if offset == 0:
                    #offset = 3 + 7
                #elif offset:
                    #offset = 0
            offset = 0
            if PersistentParameterReset:
                #sendRequest = 0
                #sendRequestCnt = 0
                txd_data[1] = 3   
                PersistentParameterReset = 0    
            elif disable15MSend:
                #sendRequest = 0
                #sendRequestCnt = 0
                txd_data[1] = 2  
                txd_data[2] = checkButton["Disable 15M"].var.get()
                disable15MSend = 0
            elif sendEnable:
                sendRequest = 0
                
                #0
                #entryVar['augerCurMax'].set((ss2[1] << 8) | ss2[0])  
                #entryVar['augerCurMin'].set((ss2[3] << 8) | ss2[2])
                #entryVar['augerCurMinEn'].set(ss2[4])
                #entryVar['augerRunOnSb'].set(ss2[5])
                #entryVar['augerTDRunOnSb'].set(ss2[6])
                txd_data[2] = (int(entryVar['augerCurMax'].get()) & 0xff)
                txd_data[3] = (int(entryVar['augerCurMax'].get()) >> 8) & 0xff
                txd_data[4] = (int(entryVar['augerCurMin'].get()) & 0xff)
                txd_data[5] = (int(entryVar['augerCurMin'].get()) >> 8) & 0xff            
                txd_data[6] = (int(entryVar['augerCurMinEn'].get()) & 0xff)
                txd_data[7] = (int(entryVar['augerRunOnSb'].get()) & 0xff)           
                txd_data[8] = (int(entryVar['augerTDRunOnSb'].get()) & 0xff) 
                
                ##5
                ##entryVar['binSignalLowHigh'].set(ss2[7])              
                ##entryVar['callWaterToutSw'].set((ss2[9] << 8) | ss2[8])
                ##entryVar['cleaningBits'].set(ss2[10])
                ##entryVar['cleanTimeMb'].set(ss2[11])
                ##entryVar['compServiceTHw'].set((ss2[13] << 8) | ss2[12])
                txd_data[9] = (int(entryVar['binSignalLowHigh'].get()) & 0xff)
                txd_data[10] = (int(entryVar['callWaterToutSw'].get())) & 0xff
                txd_data[11] = (int(entryVar['callWaterToutSw'].get()) >> 8) & 0xff
                txd_data[12] = (int(entryVar['cleaningBits'].get()) & 0xff)    
                txd_data[13] = (int(entryVar['cleanTimeMb'].get()) & 0xff)  
                txd_data[14] = (int(entryVar['compServiceTHw'].get())) & 0xff
                txd_data[15] = (int(entryVar['compServiceTHw'].get()) >> 8)  & 0xff                  
                
                ##10
                #entryVar['compStartupDly'].set((ss2[15] << 8) | ss2[14])              
                #entryVar['dlyAfterPourSb'].set(ss2[16])
                #entryVar['dlyBeforeEmptySb'].set(ss2[17])
                #entryVar['dlyBeforeRinseSb'].set(ss2[18])
                #entryVar['dip'].set(ss2[19])
                txd_data[16] = (int(entryVar['compStartupDly'].get())) & 0xff
                txd_data[17] = (int(entryVar['compStartupDly'].get()) >> 8) & 0xff
                txd_data[18] = (int(entryVar['dlyAfterPourSb'].get()) & 0xff)    
                txd_data[19] = (int(entryVar['dlyBeforeEmptySb'].get()) & 0xff)  
                txd_data[20] = (int(entryVar['dlyBeforeRinseSb'].get()) & 0xff)
                txd_data[21] = (int(entryVar['dip'].get()) & 0xff)                        
                ##15
                #entryVar['dispIMDlyDur1Tw'].set((ss2[21] << 8) | ss2[20])              
                #entryVar['dispIMDlyDur2Tw'].set((ss2[23] << 8) | ss2[22])
                #entryVar['dispIMDlyDur3Tw'].set((ss2[25] << 8) | ss2[24])
                #entryVar['dispIMDlyDur4Tw'].set((ss2[27] << 8) | ss2[26])
                #entryVar['callW.dlyDurMw'].set((ss2[29] << 8) | ss2[28]) 
                txd_data[22] = (int(entryVar['dispIMDlyDur1Tw'].get()) & 0xff)
                txd_data[23] = (int(entryVar['dispIMDlyDur1Tw'].get()) >> 8) & 0xff
                txd_data[24] = (int(entryVar['dispIMDlyDur2Tw'].get()) & 0xff)
                txd_data[25] = (int(entryVar['dispIMDlyDur2Tw'].get()) >> 8) & 0xff
                txd_data[26] = (int(entryVar['dispIMDlyDur3Tw'].get()) & 0xff)
                txd_data[27] = (int(entryVar['dispIMDlyDur3Tw'].get()) >> 8) & 0xff
                txd_data[28] = (int(entryVar['dispIMDlyDur4Tw'].get()) & 0xff)
                txd_data[29] = (int(entryVar['dispIMDlyDur4Tw'].get()) >> 8) & 0xff
                txd_data[30] = (int(entryVar['callW.dlyDurMw'].get()) & 0xff)
                txd_data[31] = (int(entryVar['callW.dlyDurMw'].get()) >> 8) & 0xff            
                ##20
                #entryVar['callW.dlyMonHb'].set(ss2[30])
                #entryVar['callW.dis'].set(ss2[31])
                #entryVar['hAmps.dlyDurMw'].set((ss2[33] << 8) | ss2[32])
                #entryVar['hAmps.dlyMonHb'].set(ss2[34])   
                #entryVar['hAmps.dis'].set(ss2[35]) 
                txd_data[32] = (int(entryVar['callW.dlyMonHb'].get()) & 0xff)    
                txd_data[33] = (int(entryVar['callW.dis'].get()) & 0xff)  
                txd_data[34] = (int(entryVar['hAmps.dlyDurMw'].get()) & 0xff)
                txd_data[35] = (int(entryVar['hAmps.dlyDurMw'].get()) >> 8) & 0xff            
                txd_data[36] = (int(entryVar['hAmps.dlyMonHb'].get()) & 0xff)
                txd_data[37] = (int(entryVar['hAmps.dis'].get()) & 0xff)                   
                ##25
                #entryVar['hPres.dlyDurMw'].set((ss2[37] << 8) | ss2[36])
                #entryVar['hPres.dlyMonHb'].set(ss2[38])   
                #entryVar['hPres.dis'].set(ss2[39])                 
                #entryVar['lowW.dlyDurMw'].set((ss2[41] << 8) | ss2[40])
                #entryVar['lowW.dlyMonHb'].set(ss2[42])
                txd_data[38] = (int(entryVar['hPres.dlyDurMw'].get()) & 0xff)
                txd_data[39] = (int(entryVar['hPres.dlyDurMw'].get()) >> 8) & 0xff            
                txd_data[40] = (int(entryVar['hPres.dlyMonHb'].get()) & 0xff)
                txd_data[41] = (int(entryVar['hPres.dis'].get()) & 0xff) 
                txd_data[42] = (int(entryVar['lowW.dlyDurMw'].get()) & 0xff)
                txd_data[43] = (int(entryVar['lowW.dlyDurMw'].get()) >> 8) & 0xff                   
                txd_data[44] = (int(entryVar['lowW.dlyMonHb'].get()) & 0xff)  
                          
                ##30
                #entryVar['lowW.dis'].set(ss2[43])
                #entryVar['wLeak.dlyDurMw'].set((ss2[45] << 8) | ss2[44])   
                #entryVar['wLeak.dlyMonHb'].set(ss2[46])                 
                #entryVar['wLeak.dis'].set(ss2[47])  
                #entryVar['drainToutSb'].set(ss2[48])
                txd_data[45] = (int(entryVar['lowW.dis'].get()) & 0xff)
                txd_data[46] = (int(entryVar['wLeak.dlyDurMw'].get()) & 0xff)
                txd_data[47] = (int(entryVar['wLeak.dlyDurMw'].get()) >> 8) & 0xff  
                txd_data[48] = (int(entryVar['wLeak.dlyMonHb'].get()) & 0xff)  
                txd_data[49] = (int(entryVar['wLeak.dis'].get()) & 0xff)
                txd_data[50] = (int(entryVar['drainToutSb'].get()) & 0xff)  
                #35
                #entryVar['fanRunOnMb'].set(ss2[49])
                #entryVar['fanStartupDly'].set((ss2[51] << 8) | ss2[50])
                #entryVar['gearSensingDly'].set((ss2[53] << 8) | ss2[52])
                #entryVar['gearStartupDly'].set((ss2[55] << 8) | ss2[54])
                #entryVar['have'].set(ss2[56])
                txd_data[51] = (int(entryVar['fanRunOnMb'].get()) & 0xff)
                txd_data[52] = (int(entryVar['fanStartupDly'].get()) & 0xff)
                txd_data[53] = (int(entryVar['fanStartupDly'].get()) >> 8) & 0xff
                txd_data[54] = (int(entryVar['gearSensingDly'].get()) & 0xff)
                txd_data[55] = (int(entryVar['gearSensingDly'].get()) >> 8) & 0xff
                txd_data[56] = (int(entryVar['gearStartupDly'].get()) & 0xff)
                txd_data[57] = (int(entryVar['gearStartupDly'].get()) >> 8) & 0xff
                txd_data[58] = (int(entryVar['have'].get()) & 0xff)
                #40          
                #entryVar['hideUItext'].set(ss2[57])
                #entryVar['imStartupDly'].set(ss2[58])
                #entryVar['lowWDlySb'].set(ss2[59])
                #entryVar['maxOffTMw'].set((ss2[61] << 8) | ss2[60])
                #entryVar['model'].set(ss2[62])  
                txd_data[59] = (int(entryVar['hideUItext'].get()) & 0xff)    
                txd_data[60] = (int(entryVar['imStartupDly'].get()) & 0xff)  
                txd_data[61] = (int(entryVar['lowWDlySb'].get()) & 0xff)
                txd_data[62] = (int(entryVar['maxOffTMw'].get()) & 0xff)
                txd_data[63] = (int(entryVar['maxOffTMw'].get()) >> 8) & 0xff            
                txd_data[64] = (int(entryVar['model'].get()) & 0xff)                      
                ##45
                #entryVar['modelInd'].set(ss2[63])
                #entryVar['pDebounceBits'].set(ss2[64])
                #entryVar['selfFlushBits1'].set(ss2[65])
                #entryVar['selfFlushBits2'].set(ss2[66])    
                #entryVar['selfFlushDur1Sw'].set((ss2[68] << 8) | ss2[67])
                txd_data[65] = (int(entryVar['modelInd'].get()) & 0xff)    
                txd_data[66] = (int(entryVar['pDebounceBits'].get()) & 0xff)  
                txd_data[67] = (int(entryVar['selfFlushBits1'].get()) & 0xff) 
                txd_data[68] = (int(entryVar['selfFlushBits2'].get()) & 0xff)   
                txd_data[69] = (int(entryVar['selfFlushDur1Sw'].get()) & 0xff)
                txd_data[70] = (int(entryVar['selfFlushDur1Sw'].get()) >> 8) & 0xff                       
                ##50
                #entryVar['selfFlushDur2Sw'].set((ss2[70] << 8) | ss2[69])
                #entryVar['selfFlushFreq1Sw'].set((ss2[72] << 8) | ss2[71])
                #entryVar['selfFlushFreq2Sw'].set((ss2[74] << 8) | ss2[73])
                #entryVar['selfFlushRepeat1'].set(ss2[75])
                #entryVar['selfFlushRepeat2'].set(ss2[76])
                txd_data[71] = (int(entryVar['selfFlushDur2Sw'].get()) & 0xff)
                txd_data[72] = (int(entryVar['selfFlushDur2Sw'].get()) >> 8) & 0xff
                txd_data[73] = (int(entryVar['selfFlushFreq1Sw'].get()) & 0xff)
                txd_data[74] = (int(entryVar['selfFlushFreq1Sw'].get()) >> 8) & 0xff
                txd_data[75] = (int(entryVar['selfFlushFreq2Sw'].get()) & 0xff)
                txd_data[76] = (int(entryVar['selfFlushFreq2Sw'].get()) >> 8) & 0xff
                txd_data[77] = (int(entryVar['selfFlushRepeat1'].get()) & 0xff)
                txd_data[78] = (int(entryVar['selfFlushRepeat2'].get()) & 0xff) 
                ##55
                #entryVar['sleepDelayMb'].set(ss2[77])
                #entryVar['lightUIDurSleep'].set(ss2[78])
                #entryVar['tdlyDrainToutSb'].set(ss2[79])
                #entryVar['tDlyFlushDip'].set(ss2[80])
                #entryVar['tDlyFlushDurSb'].set(ss2[81])
                txd_data[79] = (int(entryVar['sleepDelayMb'].get()) & 0xff)    
                txd_data[80] = (int(entryVar['lightUIDurSleep'].get()) & 0xff)  
                txd_data[81] = (int(entryVar['tdlyDrainToutSb'].get()) & 0xff) 
                txd_data[82] = (int(entryVar['tDlyFlushDip'].get()) & 0xff)   
                txd_data[83] = (int(entryVar['tDlyFlushDurSb'].get()) & 0xff)           
                ##60
                #entryVar['tDlyFlushEn'].set(ss2[82])
                #entryVar['tDlyFlushFreqMb'].set(ss2[83])
                #entryVar['tDlyFlushRep'].set(ss2[84])
                #entryVar['tDlyLongMb'].set(ss2[85])
                #entryVar['tDlyShortMb'].set(ss2[86])
                txd_data[84] = (int(entryVar['tDlyFlushEn'].get()) & 0xff)    
                txd_data[85] = (int(entryVar['tDlyFlushFreqMb'].get()) & 0xff)  
                txd_data[86] = (int(entryVar['tDlyFlushRep'].get()) & 0xff) 
                txd_data[87] = (int(entryVar['tDlyLongMb'].get()) & 0xff)   
                txd_data[88] = (int(entryVar['tDlyShortMb'].get()) & 0xff)                    
                ##65
                #entryVar['FailsafeWhenDisp'].set(ss2[87])
                #entryVar['waterManage'].set(ss2[88])                    
                #entryVar['wMDrainRetryMw'].set((ss2[90] << 8) | ss2[89])
                #entryVar['waterFillToutSw'].set((ss2[92] << 8) | ss2[91])         
                txd_data[89] = (int(entryVar['FailsafeWhenDisp'].get()) & 0xff)
                txd_data[90] = (int(entryVar['waterManage'].get()) & 0xff)            
                txd_data[91] = (int(entryVar['wMDrainRetryMw'].get()) & 0xff)
                txd_data[92] = (int(entryVar['wMDrainRetryMw'].get()) >> 8) & 0xff
                txd_data[93] = (int(entryVar['waterFillToutSw'].get()) & 0xff)
                txd_data[94] = (int(entryVar['waterFillToutSw'].get()) >> 8) & 0xff                
                txd_data[95] = (int(entryVar['modbusAddress'].get()) & 0xff)                         
                txd_data[96] = int(entryVar['noWaterADLo'].get())
                txd_data[97] = int(entryVar['noWaterADHi'].get())
                txd_data[98] = int(entryVar['noWaterADDrip'].get())          
                txd_data[99] = int(entryVar['noWaterADLeak'].get())  
                if offset == 0:
                    txd_data[100] = int(entryVar['shuttleFlushEn'].get())  
                    txd_data[101] = (int(entryVar['callWaterToutOnStartSw'].get()) & 0xff)
                    txd_data[102] = (int(entryVar['callWaterToutOnStartSw'].get()) >> 8) & 0xff 
                    i = 103
                    byte_str = [0] * 20
                    byte_str = bytes(entryVar['sn'].get(), 'ascii')
                    txd_data[i] = byte_str[0]
                    print(txd_data[i])
                    i += 1
                    txd_data[i] = byte_str[1]
                    print(txd_data[i])
                    i += 1 
                    txd_data[i] = byte_str[2]
                    i += 1                    
                    txd_data[i] = byte_str[3]
                    i += 1 
                    txd_data[i] = byte_str[4]
                    i += 1
                    txd_data[i] = byte_str[5]
                    i += 1
                    txd_data[i] = 0
                    i += 1
                    txd_data[i] = int(entryVar['spRunOnSb'].get())
                    i += 1
                    txd_data[i] = (int(entryVar['spStartupDly'].get()) & 0xff) 
                    i += 1
                    txd_data[i] = (int(entryVar['spStartupDly'].get()) >> 8) & 0xff 
                    i += 1
                    txd_data[i] = int(entryVar['icePerRes'].get()) 
                    i += 1
                    txd_data[i] = int(entryVar['DlyWUsageSb'].get())  
                    i += 1
                    #byte_str = [0] * 20
                    byte_str = bytes(entryVar['PhDModel'].get(), 'ascii')
                    txd_data[i] = byte_str[0]
                    #print(txd_data[i])
                    i += 1
                    txd_data[i] = byte_str[1]
                    #print(txd_data[i])
                    i += 1 
                    txd_data[i] = byte_str[2]
                    i += 1                    
                    txd_data[i] = byte_str[3]
                    i += 1
                    txd_data[i] = 0   
                    i += 1
                    txd_data[i] = int(entryVar['OverCurrentCnt'].get()) 

                   
                    #print(txd_data[i])
                    
                    #listtmp = [0] * 20
                    #listtmp = (list(entryVar['sn'].get()))
                    #txd_data[i] = int(listtmp[0])
                    #print(txd_data[i])
                    #i += 1
                    #txd_data[i] = listtmp[1]
                    #print(txd_data[i])
                    #i += 1 
                    #txd_data[i] = listtmp[2]
                    #i += 1                    
                    #txd_data[i] = listtmp[3]
                    #i += 1 
                    #txd_data[i] = listtmp[4]
                    #i += 1
                    #txd_data[i] = listtmp[5]
                    #print(txd_data[i])                    
                                
                txd_data[1] = 0x01 
                if sendEnable == 2:
                    txd_data[1] = 0x04 #programming
                elif sendEnable == 3:
                    txd_data[1] = 0x05 #default
                    tmp = (int(entryVar['Software Model'].get()) & 0xff)
                    txd_data[64] = tmp #!
                    txd_data[65] = tmp                   
                sendEnable = 0
            elif sendRequest:
                sendRequest = 0
                #sendRequestCnt = 0
                txd_data[1] = 6                     
                                     
            checksum = 0

            #for i in range(maxByte - 1):
                #checksum = checksum + txd_data[i]
            txd_data_t = txd_data[0:(len(txd_data) - offset)]
            
            #CRC_Calculate_txd_data(txd_data_t, offset)
        
            CRC = 0xFFFF
            #for i in range(0, CPBYTELength + 4 - 2 - offset):
            for i in range(0, len(txd_data_t) - 2):
                CRC ^= txd_data_t[i]
                for j in range(0, 8):
                    if (CRC & 0x01):
                        CRC= (CRC >> 1) ^ 0x8408
                    else:
                        CRC >>= 0x01
            CRC_txd = (~CRC) & 0xffff            
            
            #txd_data_t[maxByte + offset - 3 - 2] = (CRC_txd >> 8) & 0xff
            #txd_data_t[maxByte + offset -3 - 1] = CRC_txd & 0xff
            
            txd_data_t[len(txd_data_t) - 2] = (CRC_txd >> 8) & 0xff
            txd_data_t[len(txd_data_t) - 1] = CRC_txd & 0xff
            
            #print ([hex(x) for x in txd_data])
            #print ('[{}]'.format(', '.join(hex(x) for x in txd_data)))
            #print(maxByte)
            
            #print(txd_data_t)
            #print(len(txd_data_t))            
            
            ser.reset_input_buffer()
            try:
                #txd_data_t = txd_data[0:(len(txd_data) - offset)]
                ser.write(txd_data_t)
                #print("APP txd:")
                #print(txd_data_t)
                if txd_data_t[1] == 4:
                    
                    print(len(txd_data_t))                
            except:
                pass
                print('quit txdTask2 \r\n')
                #ser.close()
                #sys.exit()
            #if sendRequest == 0:
                #recoverCnt = 10
            #sendRequest = 0
        #checkButton["Edit"].deselect()
        time.sleep(0.25)
        sendRequest = 1
        if txd_data[1] == 4:
            time.sleep(0.25)
            checkButton["Edit"].deselect()

def rxdTask():
    global ss
    global ser
    global commCnt
    global rowIndex
    global lostCnt
    global receivingnotok
    global verificationEnable
    global CRC_rxd
    global offset
    global startup
    global cplength_offset
	
    time.sleep(0.3)

    while portOK:       
        try: 
            ss = ser.read(MAX_RECV_BYTE - offset)
        except:
            ss = b''
            #if lostCnt < 6:
                #ser.close()
                #lostCnt = 6
            pass
            #time.sleep(1)
            #print('quit rxdTask1 \r\n')
        #c = checkButton["edit"].var.get()
        c = checkButton["Edit"].var.get()
        ser.reset_input_buffer()
        if  c == 0 and ss != b'':
            ss2 = ss[2:(len(ss) - 1)]
            
            #checksum = 0 
            #for i in range(len(ss) - 1):
                #checksum = checksum + ss[i]
            #checksum = checksum & 0xff
            
            CRC_Calculate_rxd_data(len(ss))
            checksum_read = int(ss[(len(ss) - 2)] << 8 | ss[(len(ss) - 1)])
            
            #print(CRC_rxd)
            #print(checksum_read)
            print(ss)
            #print(len(ss))
            #print(len(ss2))
            #print(ss[0])
            #print(ss[(len(ss) - 1)])
            #print(checksum)
            receivingnotok = 1
            if CRC_rxd == checksum_read:# and len(ss) == MAX_RECV_BYTE: 
                #print("App rxd:")
                startup = 1
                if offset == 0:
                    cplength_offset = 0
                elif offset == 3:
                    cplength_offset = 2
                command = ss[1]
                #print(command)  
                checkButton["Receiving"].toggle()
                #print("checksum pass")
                #print(ss[67])
                #print(ss[68])
                lostCnt = 0
                receivingnotok = 0
                for columnIndex in range(0, len(ss2)):
                    rxd_data[columnIndex] = ss2[columnIndex]                  
                
                # row = 1, item
                columnIndex = 1
                for item in SaveTabs:
                    sheet.cell(row = 1, column = columnIndex).value = item
                    columnIndex = columnIndex + 1
                
                # column = 1, time    
                sheet.cell(row = rowIndex, column = 1).value = datetime.datetime.now()    
                
                #0~
                #for columnIndex in range(0, len(ss2)):
                    #sheet.cell(row = rowIndex, column = columnIndex + 2).value = rxd_data[columnIndex]       
                #print(ss2[0])
                #print(ss2[1])
                entryVar['augerCurMax'].set((ss2[1] << 8) | ss2[0])  
                entryVar['augerCurMin'].set((ss2[3] << 8) | ss2[2])
                entryVar['augerCurMinEn'].set(ss2[4])
                entryVar['augerRunOnSb'].set(ss2[5])
                entryVar['augerTDRunOnSb'].set(ss2[6])
                #5
                entryVar['binSignalLowHigh'].set(ss2[7])              
                entryVar['callWaterToutSw'].set((ss2[9] << 8) | ss2[8])
                entryVar['cleaningBits'].set(ss2[10])
                entryVar['cleanTimeMb'].set(ss2[11])
                entryVar['compServiceTHw'].set((ss2[13] << 8) | ss2[12])
                #10
                entryVar['compStartupDly'].set((ss2[15] << 8) | ss2[14])              
                entryVar['dlyAfterPourSb'].set(ss2[16])
                entryVar['dlyBeforeEmptySb'].set(ss2[17])
                entryVar['dlyBeforeRinseSb'].set(ss2[18])
                entryVar['dip'].set(ss2[19])
                #15
                entryVar['dispIMDlyDur1Tw'].set((ss2[21] << 8) | ss2[20])              
                entryVar['dispIMDlyDur2Tw'].set((ss2[23] << 8) | ss2[22])
                entryVar['dispIMDlyDur3Tw'].set((ss2[25] << 8) | ss2[24])
                entryVar['dispIMDlyDur4Tw'].set((ss2[27] << 8) | ss2[26])
                entryVar['callW.dlyDurMw'].set((ss2[29] << 8) | ss2[28])   
                #20
                entryVar['callW.dlyMonHb'].set(ss2[30])
                entryVar['callW.dis'].set(ss2[31])
                entryVar['hAmps.dlyDurMw'].set((ss2[33] << 8) | ss2[32])
                entryVar['hAmps.dlyMonHb'].set(ss2[34])   
                entryVar['hAmps.dis'].set(ss2[35]) 
                #25
                entryVar['hPres.dlyDurMw'].set((ss2[37] << 8) | ss2[36])
                entryVar['hPres.dlyMonHb'].set(ss2[38])   
                entryVar['hPres.dis'].set(ss2[39])                 
                entryVar['lowW.dlyDurMw'].set((ss2[41] << 8) | ss2[40])
                entryVar['lowW.dlyMonHb'].set(ss2[42])
                #30
                entryVar['lowW.dis'].set(ss2[43])
                entryVar['wLeak.dlyDurMw'].set((ss2[45] << 8) | ss2[44])   
                entryVar['wLeak.dlyMonHb'].set(ss2[46])                 
                entryVar['wLeak.dis'].set(ss2[47])  
                entryVar['drainToutSb'].set(ss2[48])
                #35
                entryVar['fanRunOnMb'].set(ss2[49])
                entryVar['fanStartupDly'].set((ss2[51] << 8) | ss2[50])
                entryVar['gearSensingDly'].set((ss2[53] << 8) | ss2[52])
                entryVar['gearStartupDly'].set((ss2[55] << 8) | ss2[54])
                entryVar['have'].set(ss2[56])
                #40
                entryVar['hideUItext'].set(ss2[57])
                entryVar['imStartupDly'].set(ss2[58])
                entryVar['lowWDlySb'].set(ss2[59])
                entryVar['maxOffTMw'].set((ss2[61] << 8) | ss2[60])
                entryVar['model'].set(ss2[62])   
                #45
                entryVar['modelInd'].set(ss2[63])
                entryVar['pDebounceBits'].set(ss2[64])
                entryVar['selfFlushBits1'].set(ss2[65])
                entryVar['selfFlushBits2'].set(ss2[66])    
                entryVar['selfFlushDur1Sw'].set((ss2[68] << 8) | ss2[67])
                #50
                entryVar['selfFlushDur2Sw'].set((ss2[70] << 8) | ss2[69])
                entryVar['selfFlushFreq1Sw'].set((ss2[72] << 8) | ss2[71])
                entryVar['selfFlushFreq2Sw'].set((ss2[74] << 8) | ss2[73])
                entryVar['selfFlushRepeat1'].set(ss2[75])
                entryVar['selfFlushRepeat2'].set(ss2[76])
                #55
                entryVar['sleepDelayMb'].set(ss2[77])
                entryVar['lightUIDurSleep'].set(ss2[78])
                entryVar['tdlyDrainToutSb'].set(ss2[79])
                entryVar['tDlyFlushDip'].set(ss2[80])
                entryVar['tDlyFlushDurSb'].set(ss2[81])
                #60
                entryVar['tDlyFlushEn'].set(ss2[82])
                entryVar['tDlyFlushFreqMb'].set(ss2[83])
                entryVar['tDlyFlushRep'].set(ss2[84])
                entryVar['tDlyLongMb'].set(ss2[85])
                entryVar['tDlyShortMb'].set(ss2[86])                
                #65
                entryVar['FailsafeWhenDisp'].set(ss2[87])
                entryVar['waterManage'].set(ss2[88])                    
                entryVar['wMDrainRetryMw'].set((ss2[90] << 8) | ss2[89])
                entryVar['waterFillToutSw'].set((ss2[92] << 8) | ss2[91])
                entryVar['modbusAddress'].set(ss2[93])
                # 70
                entryVar['noWaterADLo'].set(ss2[94])
                entryVar['noWaterADHi'].set(ss2[95])
                entryVar['noWaterADDrip'].set(ss2[96])
                entryVar['noWaterADLeak'].set(ss2[97])  
                if offset == 0:
                    entryVar['shuttleFlushEn'].set(ss2[98])
                #75
                    entryVar['callWaterToutOnStartSw'].set((ss2[100] << 8) | ss2[99])
                    
                    i = 101
                    tp = [None] * 6                
                
                    tp[0] = (ss2[i])
                    i += 1
                    tp[1] = (ss2[i])
                    i += 1
                    tp[2] = (ss2[i])
                    i += 1
                    tp[3] = (ss2[i])
                    i += 1
                    tp[4] = (ss2[i])
                    i += 1
                    tp[5] = (ss2[i])
                    i += 2
                    strtp = ''.join(chr(e) for e in tp)
                    #asctp = ascii(strtp)   
                    entryVar['sn'].set(strtp)
                    
                    entryVar['spRunOnSb'].set(ss2[i])
                    i += 1
                    entryVar['spStartupDly'].set((ss2[i + 1] << 8) | ss2[i])
                    i += 2
                    entryVar['icePerRes'].set(ss2[i])
                    i += 1
                    entryVar['DlyWUsageSb'].set(ss2[i])  
                    i += 1
                    tp2 = [None] * 4                
                
                    tp2[0] = (ss2[i])
                    i += 1
                    tp2[1] = (ss2[i])
                    i += 1
                    tp2[2] = (ss2[i])
                    i += 1
                    tp2[3] = (ss2[i])
                    i += 2
                    strtp2 = ''.join(chr(e) for e in tp2)
                    entryVar['PhDModel'].set(strtp2)  
                    entryVar['OverCurrentCnt'].set(ss2[i])  

                
                r_offset = 5 + 10 + 5 + 1 - offset
                entryVar['compCycCountdw'].set((ss2[101 + r_offset] << 24)  | (ss2[100 + r_offset] << 16)  | (ss2[99 + r_offset] << 8) | ss2[98 + r_offset])              
                #77
                entryVar['compOnTimeHw'].set((ss2[103 + r_offset] << 8) | ss2[102 + r_offset])
                entryVar['iDispCycCntdw'].set((ss2[107 + r_offset] << 24)  | (ss2[106 + r_offset] << 16)  | (ss2[105 + r_offset] << 8) | ss2[104 + r_offset])
                entryVar['iceDispTMdw'].set((ss2[111 + r_offset] << 24)  | (ss2[110 + r_offset] << 16)  | (ss2[109 + r_offset] << 8) | ss2[108 + r_offset])
                entryVar['powerOnTMdw'].set((ss2[115 + r_offset] << 24)  | (ss2[114 + r_offset] << 16)  | (ss2[113 + r_offset] << 8) | ss2[112 + r_offset])
                entryVar['sixMonthMdw'].set((ss2[119 + r_offset] << 24)  | (ss2[118 + r_offset] << 16)  | (ss2[117 + r_offset] << 8) | ss2[116 + r_offset])
                #82
                entryVar['wDispCycCntdw'].set((ss2[123 + r_offset] << 24)  | (ss2[122 + r_offset] << 16)  | (ss2[121 + r_offset] << 8) | ss2[120 + r_offset])
                entryVar['wDispTimeMdw'].set((ss2[127 + r_offset] << 24)  | (ss2[126 + r_offset] << 16)  | (ss2[125 + r_offset] << 8) | ss2[124 + r_offset])
                #
                entryVar['augerCurrent'].set((ss2[129 + r_offset] << 8) | ss2[128 + r_offset])
                entryVar['lowWaterSen'].set(ss2[130 + r_offset])
                entryVar['HiwWaterSen'].set(ss2[131 + r_offset])
                #87
                entryVar['dripWaterSen'].set(ss2[132 + r_offset])
                entryVar['leakWaterSen'].set(ss2[133 + r_offset])
                entryVar['binSwitch'].set(ss2[134 + r_offset])
                entryVar['tdsSwitch'].set(ss2[135 + r_offset])
                entryVar['hiPresSwitch'].set(ss2[136 + r_offset])
                #92
                entryVar['cleanSwitch'].set(ss2[137 + r_offset])
                entryVar['iceDispAux'].set(ss2[138 + r_offset])
                entryVar['maxAugerCurrent'].set((ss2[140 + r_offset] << 8) | ss2[139 + r_offset])
                entryVar['minAugerCurrent'].set((ss2[142 + r_offset] << 8) | ss2[141 + r_offset])
                #98
                entryVar['dipSwitch'].set(ss2[143 + r_offset])
                entryVar['outState'].set((ss2[147 + r_offset] << 24)  | (ss2[146 + r_offset] << 16)  | (ss2[145 + r_offset] << 8) | ss2[144 + r_offset])
                entryVar['errState'].set((ss2[149 + r_offset] << 8) | ss2[148 + r_offset])
                entryVar['myMode'].set(ss2[150 + r_offset]) 
                #102
                entryVar['version'].set((ss2[152 + r_offset] << 8) | ss2[151 + r_offset])
                entryVar['LO_CHN_AD'].set((ss2[154 + r_offset] << 8) | ss2[153 + r_offset])
                entryVar['HI_CHN_AD'].set((ss2[156 + r_offset] << 8) | ss2[155 + r_offset])
                entryVar['DRIP_CHN_AD'].set((ss2[158 + r_offset] << 8) | ss2[157 + r_offset])
                entryVar['LEAK_CHN_AD'].set((ss2[160 + r_offset] << 8) | ss2[159 + r_offset])
                #108
                entryVar['tmSinceLstFls'].set((ss2[162 + r_offset] << 8) | ss2[161 + r_offset])
                entryVar['iceCapacity'].set((ss2[164 + r_offset] << 8) | ss2[163 + r_offset])
                #output status
                value = int(entryVar['outState'].get())
                if (value & 0x01):
                    entryVar['UVLamp'].set(1)
                else:
                    entryVar['UVLamp'].set(0)
                
                if (value & 0x02):
                    entryVar['Fan'].set(1)
                else:
                    entryVar['Fan'].set(0)     
                    
                if (value & 0x04):
                    entryVar['Drain'].set(1)
                else:
                    entryVar['Drain'].set(0)   
                    
                if (value & 0x08):
                    entryVar['WaterDispenser'].set(1)
                else:
                    entryVar['WaterDispenser'].set(0)  
                    
                if (value & 0x10):
                    entryVar['FailSafe'].set(1)
                else:
                    entryVar['FailSafe'].set(0)     
                        
                if (value & 0x20):
                    entryVar['IceDispenser'].set(1)
                else:
                    entryVar['IceDispenser'].set(0)   
                        
                if (value & 0x40):
                    entryVar['Auger'].set(1)
                else:
                    entryVar['Auger'].set(0)    
                    
                if (value & 0x80):
                    entryVar['Compressor'].set(1)
                else:
                    entryVar['Compressor'].set(0)  
                
                if (value & 0x020000):
                    entryVar['ExtraOutput'].set(1)
                else:
                    entryVar['ExtraOutput'].set(0)                      
            
                  
                for i in range(CPLength, CPLength + PPLength + StatusLength): 
                    sheet.cell(row = rowIndex, column = i + 2 - CPLength).value = entryVar[RxdTabs[i]].get()
                
                for i in range(0, OutstatusLength): 
                    sheet.cell(row = rowIndex, column = i + 2 + PPLength + StatusLength).value = entryVar[OUTPUT_STATUS[i]].get()                
                rowIndex = rowIndex + 1                 
                
                #if command == 5:
                    #messagebox.showinfo('Ok',"Load Default.")
                
                if verificationEnable == 1:
                    verificationEnable = 0
                    verificationTask()
        time.sleep(0.001)  

#关于串口的任务                     
def PortTask():
    global top
    global portOK
    global ser
    global Reget
    global comNow
    global chechstatus
    global lostCnt
    global lost
    
    putCheckbutton(0, 2, "Port Open")
    Reget = 0
    baudRate = '2400'
    
    while True: 
        if Reget <= 1:      
            try: 
                ports = list(serial.tools.list_ports.comports())  #调用 serial.tools.list_ports.comports() 函数来列出所有可用的串行端口。
                    
            except:
                print ("Port dosen't exist")
                messagebox.showinfo( "!", "Port dosen't exist")
                #sys.exit() 
            portOK = 0
            mylist = [0,0,0,0]  
            j = 0 
            ok = 0
            for port, b, c, in ports:
                ok = 1
                print(ok)
                if port:
                    if Reget == 0:
                        varBaud = StringVar(top)
                        baudMenu = OptionMenu(top, varBaud, '300', '600', '1200', '2400', '4800', '9600', '14400')
                        baudMenu.grid(row = 0, column = 1)                               
                        baudRate = '14400'
                    
                    varBaud.set(baudRate) 
                    print(baudRate)                        
                        
                    mylist[j] = port
                    var = StringVar(top)
                    if j == 0:
                        mymenu = OptionMenu(top, var, mylist[0])
                    elif j == 1:
                        mymenu = OptionMenu(top, var, mylist[0], mylist[1]) 
                    elif j == 2:
                        mymenu = OptionMenu(top, var, mylist[0], mylist[1], mylist[2])  
                    else:
                        mymenu = OptionMenu(top, var, mylist[0], mylist[1], mylist[2], mylist[3])             
                    j = j + 1    
                    mymenu.grid(row = 0, column = 0)
                    if Reget == 0:
                        comNow = mylist[0]
                    Reget = 2
                    var.set(comNow)
                    print(mylist)                 
         
        
        chechstatus = checkButton["Port Open"].var.get()
        #if lostCnt >= 6:
            #lostCnt = 0
            #if portOK:         
                #ser.close()
                #portOK = 0
                #Reget = 1  
                #lost = 1
        #el
        if chechstatus: 
            if portOK == 0 and ok:   
                comNow = var.get()
                baudRate = varBaud.get()               
                try:
                    ser = serial.Serial(var.get(), baudRate, timeout = 2)#0.2
                    print(ser)
                    portOK = 1
                    rxdtd = _thread.start_new_thread(rxdTask, ())
                    txdth = _thread.start_new_thread(txdTask, ())
                except:
                    print ("Port dosen't exist or be used by other application2!")
                    #if lostCnt < 6:
                    if lost == 0:
                        messagebox.showinfo( "!", "Port dosen't exist or be used by other application2") 
                        time.sleep(5)
                lost = 0
        else:
            if portOK:
                #rxdtd.join()
                #txdth.join()                
                ser.close()
                portOK = 0
                Reget = 1
                        
        time.sleep(0.5)

def t05sTask():
    global lostCnt
    global receivingnotok
    global recoverCnt
    global sendRequest
    global verificationEnable
    global sendRequestCnt
    lostCnt = 0
    sendRequestCnt = 0
    while True:
        #if receivingnotok:
            #receivingnotok = 0
            #if lostCnt < 6:
                #lostCnt = lostCnt + 1
                #if lostCnt >= 3:
                    #lostCnt = 0
                    #ser.reset_input_buffer()
                    
        #time.sleep(0.5)
        if verificationEnable > 1:
            verificationEnable -= 1
        
        sendRequestCnt = sendRequestCnt + 1
        if sendRequestCnt >= 5:
            sendRequestCnt = 0
            #sendRequest = 1
        
        #if recoverCnt:
            #recoverCnt = recoverCnt - 1
            #if recoverCnt == 0:
                #checkButton["Edit"].deselect()
        
        if portOK:
            lostCnt = lostCnt + 1
            if lostCnt >= 30:
                lostCnt = 0
                ser.reset_input_buffer()
    
        time.sleep(0.05)        
        #print("lostCnt = ", lostCnt)

#********MAIN*******  
portOK = 0
top = tkinter.Tk()
top.title("Follett Ice Machine")

_thread.start_new_thread(topTask, ())
_thread.start_new_thread(excelTask, ())
_thread.start_new_thread(PortTask, ())
_thread.start_new_thread(SaveTask, ())
_thread.start_new_thread(SaveCPTask, ()) 
#_thread.start_new_thread(LoadCPTask, ())   
_thread.start_new_thread(t05sTask, ())
#_thread.start_new_thread(rxdTask, ())  

# Code to add widgets will go here...

top.mainloop()
